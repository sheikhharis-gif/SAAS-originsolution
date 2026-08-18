"""
Lead discovery engine — multi-source business lead scraper.

Ported from the standalone LeadGen Pro tool (`lead generation/complete_lead_gen.py`)
so the Sales portal can run the same search/scrape/score pipeline in-process.
Scrapes Google Search (+ optional Google Places API), Facebook and LinkedIn
pages for businesses matching a city/country/keyword search, visits each
result to extract emails/phones/socials, then scores and qualifies each lead.
"""

import csv
import json
import os
import re
import time
import random
import hashlib
import urllib.parse
from datetime import datetime
from typing import List, Dict, Optional, Set, Tuple
from collections import defaultdict

import requests
from bs4 import BeautifulSoup

# ============================================================
# CONFIGURATION
# ============================================================
SEARCH_DELAY_MIN = 1.5
SEARCH_DELAY_MAX = 3.0
WEBSITE_TIMEOUT = 4
MAX_CONCURRENT_VISITS = 5
MAX_RESULTS_PER_KEYWORD = 50
CONTACT_PAGES = ["/contact", "/about"]

# ============================================================
# DATA STORAGE
# ============================================================
all_leads: List[Dict] = []
search_history: List[Dict] = []
search_logs: List[Dict] = []

# ============================================================
# USER AGENT POOL
# ============================================================
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:126.0) Gecko/20100101 Firefox/126.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14.5; rv:126.0) Gecko/20100101 Firefox/126.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_5) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
]

# ============================================================
# BUSINESS CATEGORIES
# ============================================================
BUSINESS_CATEGORIES = {
    "Food & Dining": ["restaurants", "pizza places", "coffee shops", "bakery", "ice cream shops", "bars", "breweries"],
    "Beauty & Wellness": ["hair salons", "barbershops", "nail salons", "spas", "massage therapy"],
    "Health & Medical": ["dentists", "dental clinics", "doctors", "pediatricians", "dermatologists", "optometrists"],
    "Professional Services": ["lawyers", "accountants", "real estate agents", "financial advisors", "insurance agents"],
    "Home Services": ["plumbers", "electricians", "contractors", "handyman", "roofers", "painters", "landscapers"],
    "Automotive": ["auto repair", "mechanics", "car dealerships", "auto body shops", "car washes", "tire shops"],
    "Fitness": ["gyms", "fitness centers", "yoga studios", "personal trainers", "crossfit", "martial arts"],
    "Retail": ["clothing stores", "electronics stores", "furniture stores", "bookstores", "pharmacies"],
    "Education": ["tutoring", "music lessons", "dance classes", "art classes", "language schools"],
    "Pets": ["veterinarians", "pet grooming", "pet boarding", "pet stores"],
    "Events": ["event planners", "wedding planners", "photographers", "videographers", "caterers"],
    "Hotels & Travel": ["hotels", "motels", "bed and breakfasts", "vacation rentals"],
    "Cleaning": ["house cleaning", "carpet cleaning", "window cleaning", "commercial cleaning"],
}

# ============================================================
# LEAD SCORING RULES
# ============================================================
LEAD_SCORING_RULES = {
    "has_email": 15,
    "has_phone": 12,
    "has_website": 10,
    "has_social_media": 8,
    "has_multiple_contacts": 10,
    "rating_high": 15,
    "many_reviews": 10,
    "complete_info": 10,
    "category_bonus": 10,
}


# ============================================================
# HELPER UTILITIES
# ============================================================
def _add_log(message: str, level: str = "info"):
    """Append to the global log list."""
    search_logs.append({
        "timestamp": datetime.now().strftime("%H:%M:%S"),
        "message": message,
        "level": level,
    })
    icon = {"info": "[INFO]", "success": "[OK]", "warning": "[WARN]", "error": "[ERR]"}.get(level, "[LOG]")
    try:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {icon} {message}")
    except UnicodeEncodeError:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {icon} {message.encode('ascii', 'replace').decode()}")


def _extract_business_name(title: str, domain: str) -> str:
    """Try to derive a clean business name from page title."""
    if not title:
        return domain.split(".")[0].title()
    for sep in [" | ", " - ", " — ", " – ", " :: ", " » "]:
        if sep in title:
            title = title.split(sep)[0]
    return title.strip()


def _guess_category(title: str, snippet: str) -> str:
    """Guess business category from title and snippet text."""
    combined = (title + " " + snippet).lower()
    category_keywords = {
        "Restaurant": ["restaurant", "cafe", "bakery", "pizza", "sushi", "coffee", "dining", "food"],
        "Salon": ["salon", "barber", "hair", "nails", "beauty"],
        "Spa": ["spa", "massage", "wellness"],
        "Dental": ["dentist", "dental", "orthodont"],
        "Medical": ["doctor", "medical", "clinic", "physician", "health"],
        "Legal": ["lawyer", "attorney", "law firm", "legal"],
        "Real Estate": ["real estate", "realtor", "realty", "property"],
        "Accounting": ["accountant", "accounting", "cpa", "tax"],
        "Plumbing": ["plumber", "plumbing", "drain"],
        "Electrical": ["electrician", "electrical", "wiring"],
        "Construction": ["contractor", "construction", "remodel", "renovation"],
        "Auto Repair": ["auto repair", "mechanic", "car repair", "auto body"],
        "Gym": ["gym", "fitness", "workout", "crossfit"],
        "Hotel": ["hotel", "motel", "inn", "lodge", "resort"],
        "Retail": ["store", "shop", "boutique", "retail"],
        "Cleaning": ["cleaning", "janitorial", "maid service"],
        "Photography": ["photographer", "photography", "photo studio"],
        "Education": ["tutor", "school", "academy", "lesson", "training"],
        "Veterinary": ["vet", "veterinary", "animal hospital", "pet"],
        "Insurance": ["insurance", "coverage", "policy"],
    }
    for cat, kws in category_keywords.items():
        for kw in kws:
            if kw in combined:
                return cat
    return "Business"


# ============================================================
# GOOGLE SEARCH SCRAPER
# ============================================================
class GoogleSearchScraper:
    """Multi-engine Web Search Scraper (DuckDuckGo HTML + Google Search fallback)."""

    def __init__(self):
        self._request_count = 0

    def _get_headers(self) -> Dict[str, str]:
        ua = random.choice(USER_AGENTS)
        return {
            "User-Agent": ua,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
        }

    def search(self, query: str, num_results: int = 30) -> Dict[str, List[Dict]]:
        """
        Search Web for organic results, Facebook pages, and LinkedIn profiles.
        """
        websites: List[Dict] = []
        facebook: List[Dict] = []
        linkedin: List[Dict] = []
        seen_urls: Set[str] = set()

        # 1. Fetch Organic Web Results via DuckDuckGo HTML API
        ddg_results = self._search_ddg(query)
        for r in ddg_results:
            if r["url"] in seen_urls:
                continue
            seen_urls.add(r["url"])
            if r["type"] == "facebook":
                facebook.append(r)
            elif r["type"] == "linkedin":
                linkedin.append(r)
            else:
                websites.append(r)

        # 2. Specifically query Facebook pages if few found
        if len(facebook) < 3:
            fb_results = self._search_ddg(f"site:facebook.com {query}")
            for r in fb_results:
                if r["url"] not in seen_urls and r["type"] == "facebook":
                    seen_urls.add(r["url"])
                    facebook.append(r)

        # 3. Specifically query LinkedIn company pages if few found
        if len(linkedin) < 3:
            li_results = self._search_ddg(f"site:linkedin.com/company {query}")
            for r in li_results:
                if r["url"] not in seen_urls and r["type"] == "linkedin":
                    seen_urls.add(r["url"])
                    linkedin.append(r)

        # 4. Fallback to Google Search if DDG returns insufficient results
        if len(websites) < 3:
            google_results = self._search_google(query)
            for r in google_results:
                if r["url"] not in seen_urls:
                    seen_urls.add(r["url"])
                    if r["type"] == "facebook":
                        facebook.append(r)
                    elif r["type"] == "linkedin":
                        linkedin.append(r)
                    else:
                        websites.append(r)

        return {
            "websites": websites[:num_results],
            "facebook": facebook[:15],
            "linkedin": linkedin[:15],
        }

    def _search_ddg(self, query: str) -> List[Dict]:
        """Query DuckDuckGo HTML API and extract results."""
        results: List[Dict] = []
        url = "https://html.duckduckgo.com/html/"
        data = {"q": query}
        skip_domains = [
            "duckduckgo.com", "google.com", "youtube.com", "wikipedia.org",
            "pinterest.com", "reddit.com", "yelp.com", "yellowpages.com",
            "bbb.org", "tripadvisor.com", "consumeraffairs.com", "angi.com",
            "expertise.com", "thumbtack.com", "trustpilot.com",
        ]

        try:
            resp = requests.post(url, data=data, headers=self._get_headers(), timeout=8)
            self._request_count += 1
            if resp.status_code == 200:
                soup = BeautifulSoup(resp.text, "html.parser")
                for div in soup.select("div.result"):
                    a_url = div.select_one("a.result__url")
                    a_title = div.select_one("a.result__a")
                    a_snippet = div.select_one("a.result__snippet")

                    if not a_url or not a_url.get("href"):
                        continue

                    raw_href = a_url.get("href", "").strip()
                    m = re.search(r"uddg=([^&]+)", raw_href)
                    actual_url = urllib.parse.unquote(m.group(1)) if m else raw_href

                    if not actual_url.startswith("http"):
                        actual_url = "https://" + actual_url.lstrip("/")

                    domain = urllib.parse.urlparse(actual_url).netloc.lower().replace("www.", "")

                    if any(sd in domain for sd in skip_domains):
                        continue

                    title = a_title.get_text(strip=True) if a_title else ""
                    snippet = a_snippet.get_text(strip=True) if a_snippet else ""

                    res_type = "website"
                    if "facebook.com" in domain:
                        res_type = "facebook"
                    elif "linkedin.com" in domain:
                        res_type = "linkedin"

                    results.append({
                        "title": title,
                        "url": actual_url,
                        "snippet": snippet,
                        "domain": domain,
                        "type": res_type,
                    })

        except Exception as e:
            _add_log(f"Search provider note: {str(e)[:60]}", "warning")

        return results

    def _search_google(self, query: str) -> List[Dict]:
        """Fallback Google Search HTML parser."""
        results: List[Dict] = []
        params = {"q": query, "num": "15", "hl": "en"}
        url = "https://www.google.com/search?" + urllib.parse.urlencode(params)
        skip_domains = [
            "google.com", "google.co", "youtube.com", "wikipedia.org",
            "pinterest.com", "reddit.com", "yelp.com", "yellowpages.com",
            "bbb.org", "tripadvisor.com",
        ]

        try:
            resp = requests.get(url, headers=self._get_headers(), timeout=8)
            if resp.status_code == 200:
                soup = BeautifulSoup(resp.text, "html.parser")
                for div in soup.select("div.g, div.MjjYud"):
                    link_el = div.select_one("a[href]")
                    if not link_el:
                        continue
                    href = link_el.get("href", "")
                    if not href.startswith("http"):
                        continue
                    domain = urllib.parse.urlparse(href).netloc.lower().replace("www.", "")
                    if any(sd in domain for sd in skip_domains):
                        continue
                    title_el = div.select_one("h3")
                    title = title_el.get_text(strip=True) if title_el else ""
                    snippet_el = div.select_one("div.VwiC3b, span.aCOpRe")
                    snippet = snippet_el.get_text(strip=True) if snippet_el else ""

                    res_type = "website"
                    if "facebook.com" in domain:
                        res_type = "facebook"
                    elif "linkedin.com" in domain:
                        res_type = "linkedin"

                    results.append({
                        "title": title,
                        "url": href,
                        "snippet": snippet,
                        "domain": domain,
                        "type": res_type,
                    })
        except Exception:
            pass

        return results


# ============================================================
# GOOGLE PLACES API SCRAPER
# ============================================================
class GooglePlacesScraper:
    """Uses Google Places Text Search + Details API for structured business data."""

    def __init__(self, api_key: str):
        self.api_key = api_key

    def search(self, query: str, max_results: int = 20) -> List[Dict]:
        """Search Google Places and return enriched business data."""
        results: List[Dict] = []
        next_page_token = None

        url = "https://maps.googleapis.com/maps/api/place/textsearch/json"
        params = {"query": query, "key": self.api_key}

        # Fetch up to 3 pages (60 results max from Places API)
        for page_num in range(3):
            if len(results) >= max_results:
                break

            if page_num > 0:
                if not next_page_token:
                    break
                time.sleep(2)  # Required delay for pagetoken
                params["pagetoken"] = next_page_token

            try:
                resp = requests.get(url, params=params, timeout=15)
                if resp.status_code != 200:
                    _add_log(f"Places API error: status {resp.status_code}", "error")
                    break

                data = resp.json()
                status = data.get("status", "")

                if status == "REQUEST_DENIED":
                    _add_log("Places API: Invalid API key or API not enabled", "error")
                    break
                elif status == "OVER_QUERY_LIMIT":
                    _add_log("Places API: Query limit exceeded", "warning")
                    break
                elif status not in ("OK", "ZERO_RESULTS"):
                    _add_log(f"Places API status: {status}", "warning")
                    break

                places = data.get("results", [])
                next_page_token = data.get("next_page_token")

                for place in places:
                    if len(results) >= max_results:
                        break
                    details = self._get_details(place.get("place_id", ""))
                    results.append(self._build_lead(place, details))

            except Exception as e:
                _add_log(f"Places API error: {str(e)[:100]}", "error")
                break

        return results

    def _get_details(self, place_id: str) -> Dict:
        """Get detailed place information."""
        if not place_id:
            return {}

        url = "https://maps.googleapis.com/maps/api/place/details/json"
        params = {
            "place_id": place_id,
            "key": self.api_key,
            "fields": "formatted_phone_number,website,opening_hours,url",
        }

        try:
            resp = requests.get(url, params=params, timeout=10)
            if resp.status_code == 200:
                result = resp.json().get("result", {})
                return {
                    "phone": result.get("formatted_phone_number", ""),
                    "website": result.get("website", ""),
                    "opening_hours": result.get("opening_hours", {}).get("weekday_text", []),
                    "maps_url": result.get("url", ""),
                }
        except Exception:
            pass

        return {}

    def _extract_postal_code(self, address: str) -> str:
        """Extract postal code from address string."""
        match = re.search(r"\b\d{5}(?:-\d{4})?\b", address)
        return match.group() if match else ""

    def _build_lead(self, place: Dict, details: Dict) -> Dict:
        """Build a lead dict from Places API data."""
        address = place.get("formatted_address", "")
        rating = place.get("rating", 0)
        reviews = place.get("user_ratings_total", 0)
        website = details.get("website", "")
        phone = details.get("phone", "")
        types = place.get("types", [])

        # Determine category
        type_map = {
            "restaurant": "Restaurant", "cafe": "Cafe", "bar": "Bar",
            "hair_care": "Salon", "spa": "Spa", "gym": "Gym",
            "dentist": "Dental", "doctor": "Medical", "lawyer": "Legal",
            "real_estate_agency": "Real Estate", "accounting": "Accounting",
            "plumber": "Plumbing", "electrician": "Electrical",
            "car_repair": "Auto Repair", "hotel": "Hotel", "store": "Retail",
        }
        category = "Business"
        for t in types:
            if t in type_map:
                category = type_map[t]
                break

        return {
            "business_name": place.get("name", "Unknown"),
            "phone": phone,
            "website": website,
            "address": address,
            "postal_code": self._extract_postal_code(address),
            "rating": rating,
            "reviews_count": reviews,
            "category": category,
            "maps_url": details.get("maps_url", f"https://maps.google.com/?q={place.get('place_id', '')}"),
            "lat": place.get("geometry", {}).get("location", {}).get("lat", 0),
            "lng": place.get("geometry", {}).get("location", {}).get("lng", 0),
            "source": "google_places",
        }


# ============================================================
# FACEBOOK PAGE EXTRACTOR
# ============================================================
class FacebookExtractor:
    """Extract business info from Facebook pages found in search results."""

    def extract(self, url: str, title: str = "", snippet: str = "") -> Dict:
        """
        Extract business info from a Facebook page.
        Uses page HTML + structured data extraction.
        """
        result = {
            "business_name": "",
            "phone": "",
            "email": "",
            "website": "",
            "address": "",
            "category": "",
            "facebook_url": url,
            "source": "facebook",
            "accessible": False,
        }

        # Extract business name from title
        fb_name = title
        for suffix in [" - Facebook", " | Facebook", "- Home | Facebook",
                        " - Home", " - About", " - Posts", " - Reviews"]:
            fb_name = fb_name.replace(suffix, "")
        result["business_name"] = fb_name.strip()

        # Try to extract info from snippet
        snippet_info = self._parse_snippet(snippet)
        result.update({k: v for k, v in snippet_info.items() if v})

        # Try to fetch the page
        try:
            headers = {
                "User-Agent": random.choice(USER_AGENTS),
                "Accept": "text/html,application/xhtml+xml",
                "Accept-Language": "en-US,en;q=0.9",
            }
            resp = requests.get(url, headers=headers, timeout=WEBSITE_TIMEOUT,
                              allow_redirects=True, verify=False)

            if resp.status_code == 200 and "text/html" in resp.headers.get("content-type", ""):
                result["accessible"] = True
                page_info = self._parse_page(resp.text)
                # Merge — page data overrides snippet data if non-empty
                for k, v in page_info.items():
                    if v:
                        result[k] = v

        except Exception:
            pass

        return result

    def _parse_snippet(self, snippet: str) -> Dict:
        """Extract phone/email/address from Google snippet of FB page."""
        info: Dict = {}

        # Phone patterns in snippets
        phone_match = re.search(
            r"(?:Phone|Tel|Call)[:\s]*([+\d\s\-().]{7,20})", snippet, re.IGNORECASE
        )
        if phone_match:
            info["phone"] = phone_match.group(1).strip()

        # General phone pattern
        if not info.get("phone"):
            phone_match2 = re.search(
                r"(?:\+?\d{1,3}[\s\-.]?)?\(?\d{2,4}\)?[\s\-.]?\d{3,4}[\s\-.]?\d{3,4}", snippet
            )
            if phone_match2:
                info["phone"] = phone_match2.group().strip()

        # Email
        email_match = re.search(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", snippet)
        if email_match:
            info["email"] = email_match.group()

        # Rating
        rating_match = re.search(r"(\d+\.?\d*)\s*(?:stars?|rating)", snippet, re.IGNORECASE)
        if rating_match:
            try:
                info["rating"] = float(rating_match.group(1))
            except ValueError:
                pass

        # Reviews
        reviews_match = re.search(r"(\d+[,\d]*)\s*(?:reviews?|ratings?)", snippet, re.IGNORECASE)
        if reviews_match:
            try:
                info["reviews_count"] = int(reviews_match.group(1).replace(",", ""))
            except ValueError:
                pass

        return info

    def _parse_page(self, html: str) -> Dict:
        """Extract structured data from Facebook page HTML."""
        info: Dict = {}
        soup = BeautifulSoup(html, "html.parser")

        # Phone numbers from page content
        phone_pattern = re.compile(
            r"(?:\+?\d{1,3}[\s\-.]?)?\(?\d{2,4}\)?[\s\-.]?\d{3,4}[\s\-.]?\d{3,4}"
        )
        email_pattern = re.compile(
            r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", re.IGNORECASE
        )
        website_pattern = re.compile(
            r'https?://(?!(?:www\.)?facebook\.com)[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}[/\w.-]*',
            re.IGNORECASE,
        )

        text = soup.get_text(" ", strip=True)

        # Emails (exclude facebook emails)
        emails = email_pattern.findall(text)
        emails = [e for e in emails if "facebook" not in e.lower() and "fb.com" not in e.lower()]
        if emails:
            info["email"] = emails[0]

        # Phone
        phones = phone_pattern.findall(text)
        phones = [p.strip() for p in phones if len(re.sub(r"\D", "", p)) >= 7]
        if phones:
            info["phone"] = phones[0]

        # External website
        for a_tag in soup.select("a[href]"):
            href = a_tag.get("href", "")
            if href.startswith("http") and "facebook.com" not in href and "fb.com" not in href:
                # Check if it looks like a real website
                parsed = urllib.parse.urlparse(href)
                if parsed.netloc and "." in parsed.netloc:
                    info["website"] = href
                    break

        # Try JSON-LD structured data
        for script in soup.select("script[type='application/ld+json']"):
            try:
                ld_data = json.loads(script.string or "")
                if isinstance(ld_data, dict):
                    if ld_data.get("telephone"):
                        info["phone"] = ld_data["telephone"]
                    if ld_data.get("email"):
                        info["email"] = ld_data["email"]
                    if ld_data.get("url") and "facebook.com" not in ld_data.get("url", ""):
                        info["website"] = ld_data["url"]
                    addr = ld_data.get("address", {})
                    if isinstance(addr, dict) and addr.get("streetAddress"):
                        info["address"] = f"{addr.get('streetAddress', '')}, {addr.get('addressLocality', '')}"
            except Exception:
                continue

        # Try meta tags
        for meta in soup.select("meta[property], meta[name]"):
            prop = meta.get("property", "") or meta.get("name", "")
            content = meta.get("content", "")
            if prop == "og:description" and content and not info.get("address"):
                # Often contains category and address info
                if any(kw in content.lower() for kw in ["located", "address", "street", "avenue", "road"]):
                    info["address"] = content[:200]

        return info


# ============================================================
# LINKEDIN COMPANY EXTRACTOR
# ============================================================
class LinkedInExtractor:
    """Extract company info from LinkedIn pages found in search results."""

    def extract(self, url: str, title: str = "", snippet: str = "") -> Dict:
        """Extract company info from a LinkedIn company page."""
        result = {
            "business_name": "",
            "phone": "",
            "email": "",
            "website": "",
            "address": "",
            "category": "",
            "linkedin_url": url,
            "employee_count": "",
            "source": "linkedin",
            "accessible": False,
        }

        # Extract company name from title
        li_name = title
        for suffix in [" | LinkedIn", " - LinkedIn", " on LinkedIn",
                        " - Company Profile", " Overview", " - Overview"]:
            li_name = li_name.replace(suffix, "")
        result["business_name"] = li_name.strip()

        # Parse snippet for info
        snippet_info = self._parse_snippet(snippet)
        result.update({k: v for k, v in snippet_info.items() if v})

        # Try to fetch the page
        try:
            headers = {
                "User-Agent": random.choice(USER_AGENTS),
                "Accept": "text/html,application/xhtml+xml",
                "Accept-Language": "en-US,en;q=0.9",
            }
            resp = requests.get(url, headers=headers, timeout=WEBSITE_TIMEOUT,
                              allow_redirects=True, verify=False)

            if resp.status_code == 200 and "text/html" in resp.headers.get("content-type", ""):
                result["accessible"] = True
                page_info = self._parse_page(resp.text)
                for k, v in page_info.items():
                    if v:
                        result[k] = v

        except Exception:
            pass

        return result

    def _parse_snippet(self, snippet: str) -> Dict:
        """Extract info from Google snippet of LinkedIn page."""
        info: Dict = {}

        # Employee count
        emp_match = re.search(r"(\d+[,\d]*)\s*(?:employees?|followers?|staff)", snippet, re.IGNORECASE)
        if emp_match:
            info["employee_count"] = emp_match.group(1)

        # Industry/category
        industry_match = re.search(
            r"(?:Industry|Sector)[:\s]*([A-Za-z\s&,]+?)(?:\.|;|\d|$)", snippet, re.IGNORECASE
        )
        if industry_match:
            info["category"] = industry_match.group(1).strip()

        # Location
        loc_match = re.search(
            r"(?:Location|Based in|Headquartered in|Located in)[:\s]*([^.;]+)", snippet, re.IGNORECASE
        )
        if loc_match:
            info["address"] = loc_match.group(1).strip()

        # Phone
        phone_match = re.search(
            r"(?:\+?\d{1,3}[\s\-.]?)?\(?\d{2,4}\)?[\s\-.]?\d{3,4}[\s\-.]?\d{3,4}", snippet
        )
        if phone_match:
            info["phone"] = phone_match.group().strip()

        # Website URL in snippet
        web_match = re.search(
            r"(https?://(?!linkedin\.com)[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}[/\w.-]*)", snippet
        )
        if web_match:
            info["website"] = web_match.group(1)

        return info

    def _parse_page(self, html: str) -> Dict:
        """Extract structured data from LinkedIn page HTML."""
        info: Dict = {}
        soup = BeautifulSoup(html, "html.parser")

        # Try JSON-LD
        for script in soup.select("script[type='application/ld+json']"):
            try:
                ld = json.loads(script.string or "")
                if isinstance(ld, dict):
                    if ld.get("@type") in ("Organization", "Corporation", "LocalBusiness"):
                        if ld.get("telephone"):
                            info["phone"] = ld["telephone"]
                        if ld.get("email"):
                            info["email"] = ld["email"]
                        if ld.get("url") and "linkedin.com" not in ld.get("url", ""):
                            info["website"] = ld["url"]
                        if ld.get("numberOfEmployees"):
                            emp = ld["numberOfEmployees"]
                            if isinstance(emp, dict):
                                info["employee_count"] = str(emp.get("value", ""))
                            else:
                                info["employee_count"] = str(emp)
                        addr = ld.get("address", {})
                        if isinstance(addr, dict) and addr.get("addressLocality"):
                            info["address"] = f"{addr.get('streetAddress', '')}, {addr.get('addressLocality', '')}, {addr.get('addressRegion', '')}".strip(", ")
            except Exception:
                continue

        # Try meta tags
        for meta in soup.select("meta[property], meta[name]"):
            prop = meta.get("property", "") or meta.get("name", "")
            content = meta.get("content", "")
            if not content:
                continue
            if prop == "og:title" and not info.get("business_name"):
                name = content
                for suffix in [" | LinkedIn", " - LinkedIn"]:
                    name = name.replace(suffix, "")
                info["business_name"] = name.strip()
            if prop == "og:description" and content:
                # Extract website from description
                web_match = re.search(
                    r"(https?://(?!linkedin\.com)[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})", content
                )
                if web_match and not info.get("website"):
                    info["website"] = web_match.group(1)

        # Look for external links
        email_pattern = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", re.IGNORECASE)
        text = soup.get_text(" ", strip=True)
        emails = email_pattern.findall(text)
        emails = [e for e in emails if "linkedin" not in e.lower()]
        if emails and not info.get("email"):
            info["email"] = emails[0]

        return info


# ============================================================
# CONTACT EXTRACTOR (for regular websites)
# ============================================================
class ContactExtractor:
    """Visits websites to extract emails, phone numbers, and social links."""

    EMAIL_PATTERN = re.compile(
        r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", re.IGNORECASE
    )
    PHONE_PATTERN = re.compile(
        r"(?:(?:\+?\d{1,3}[\s\-.]?)?\(?\d{2,4}\)?[\s\-.]?\d{3,4}[\s\-.]?\d{3,4})"
    )
    SOCIAL_PATTERNS = {
        "facebook": re.compile(r"https?://(?:www\.)?facebook\.com/[a-zA-Z0-9._\-/]+", re.IGNORECASE),
        "instagram": re.compile(r"https?://(?:www\.)?instagram\.com/[a-zA-Z0-9._\-/]+", re.IGNORECASE),
        "linkedin": re.compile(r"https?://(?:www\.)?linkedin\.com/(?:in|company)/[a-zA-Z0-9._\-/]+", re.IGNORECASE),
        "twitter": re.compile(r"https?://(?:www\.)?(?:twitter\.com|x\.com)/[a-zA-Z0-9._\-/]+", re.IGNORECASE),
    }
    JUNK_EMAILS = re.compile(
        r"(?:example\.com|test\.com|email\.com|sentry\.io|wixpress|"
        r"schema\.org|googleapis|w3\.org|placeholder|noreply|no-reply|"
        r"\.png|\.jpg|\.gif|\.svg|\.webp|\.css|\.js)",
        re.IGNORECASE,
    )

    def _clean_emails(self, raw_emails: List[str]) -> List[str]:
        clean: List[str] = []
        seen: Set[str] = set()
        for email in raw_emails:
            email = email.lower().strip().rstrip(".")
            if email in seen:
                continue
            if self.JUNK_EMAILS.search(email):
                continue
            if len(email) < 6 or len(email) > 254:
                continue
            parts = email.rsplit(".", 1)
            if len(parts) != 2 or len(parts[1]) < 2:
                continue
            seen.add(email)
            clean.append(email)
        return clean

    def _clean_phones(self, raw_phones: List[str]) -> List[str]:
        clean: List[str] = []
        seen: Set[str] = set()
        for phone in raw_phones:
            phone = re.sub(r"\s+", " ", phone).strip()
            digits_only = re.sub(r"\D", "", phone)
            if len(digits_only) < 7 or len(digits_only) > 15:
                continue
            if digits_only in seen:
                continue
            seen.add(digits_only)
            clean.append(phone)
        return clean

    def extract_from_html(self, html: str, base_url: str = "") -> Dict:
        emails = self._clean_emails(self.EMAIL_PATTERN.findall(html))
        phones = self._clean_phones(self.PHONE_PATTERN.findall(html))
        socials = {}
        for platform, pattern in self.SOCIAL_PATTERNS.items():
            matches = pattern.findall(html)
            if matches:
                socials[platform] = matches[0].rstrip("/")
        return {"emails": emails[:5], "phones": phones[:5], "socials": socials}

    def fetch_and_extract(self, url: str) -> Dict:
        result = {"emails": [], "phones": [], "socials": {}, "accessible": False}
        try:
            headers = {"User-Agent": random.choice(USER_AGENTS)}
            resp = requests.get(url, headers=headers, timeout=WEBSITE_TIMEOUT,
                              allow_redirects=True, verify=False)
            if resp.status_code == 200 and "text/html" in resp.headers.get("content-type", ""):
                result["accessible"] = True
                extracted = self.extract_from_html(resp.text, url)
                result["emails"] = extracted["emails"]
                result["phones"] = extracted["phones"]
                result["socials"] = extracted["socials"]

                if not result["emails"]:
                    parsed = urllib.parse.urlparse(url)
                    base = f"{parsed.scheme}://{parsed.netloc}"
                    for contact_path in CONTACT_PAGES:
                        try:
                            r2 = requests.get(base + contact_path, headers=headers,
                                            timeout=2.5, allow_redirects=True, verify=False)
                            if r2.status_code == 200 and "text/html" in r2.headers.get("content-type", ""):
                                extra = self.extract_from_html(r2.text, base + contact_path)
                                result["emails"].extend(extra["emails"])
                                result["phones"].extend(extra["phones"])
                                for k, v in extra["socials"].items():
                                    if k not in result["socials"]:
                                        result["socials"][k] = v
                                if result["emails"]:
                                    break
                        except Exception:
                            continue

                result["emails"] = list(dict.fromkeys(result["emails"]))[:5]
                result["phones"] = list(dict.fromkeys(result["phones"]))[:5]

        except Exception:
            pass
        return result


# ============================================================
# LEAD SCORER
# ============================================================
class LeadScorer:
    """Score and qualify discovered leads."""

    @staticmethod
    def score(lead: Dict) -> Tuple[int, str]:
        score = 0

        if lead.get("email"):
            score += LEAD_SCORING_RULES["has_email"]
        if lead.get("phone"):
            score += LEAD_SCORING_RULES["has_phone"]
        if lead.get("website") and lead.get("website_accessible", True):
            score += LEAD_SCORING_RULES["has_website"]

        social_count = sum(
            1 for k in ["facebook_url", "instagram_url", "linkedin_url", "twitter_url"]
            if lead.get(k)
        )
        if social_count > 0:
            score += min(LEAD_SCORING_RULES["has_social_media"], social_count * 3)

        contact_channels = sum([
            bool(lead.get("email")),
            bool(lead.get("phone")),
            social_count > 0,
            bool(lead.get("website")),
        ])
        if contact_channels >= 3:
            score += LEAD_SCORING_RULES["has_multiple_contacts"]

        rating = lead.get("rating", 0) or 0
        if rating >= 4.0:
            score += LEAD_SCORING_RULES["rating_high"]
        elif rating >= 3.0:
            score += 8

        reviews = lead.get("reviews_count", 0) or 0
        if reviews >= 50:
            score += LEAD_SCORING_RULES["many_reviews"]
        elif reviews >= 20:
            score += 5

        score = min(100, max(0, score))
        quality = "Hot" if score >= 70 else ("Warm" if score >= 40 else "Cold")
        return score, quality


# ============================================================
# SEARCH ORCHESTRATION
# ============================================================
def _build_lead_from_website(result: Dict, contacts: Dict, keyword: str,
                             city: str, country: str, scorer: LeadScorer) -> Dict:
    """Build a lead dict from a website result + extracted contacts."""
    business_name = _extract_business_name(result["title"], result["domain"])
    category = _guess_category(result["title"], result["snippet"])

    lead = {
        "Business Name": business_name,
        "Phone Number": contacts["phones"][0] if contacts["phones"] else "N/A",
        "All Phones": ", ".join(contacts["phones"]) if contacts["phones"] else "N/A",
        "Email": contacts["emails"][0] if contacts["emails"] else "N/A",
        "All Emails": ", ".join(contacts["emails"]) if contacts["emails"] else "N/A",
        "Website": result["url"],
        "Domain": result["domain"],
        "Address": f"{city}, {country}",
        "Category": category,
        "Snippet": result["snippet"][:200] if result.get("snippet") else "",
        "Facebook URL": contacts["socials"].get("facebook", ""),
        "Instagram URL": contacts["socials"].get("instagram", ""),
        "LinkedIn URL": contacts["socials"].get("linkedin", ""),
        "Twitter URL": contacts["socials"].get("twitter", ""),
        "Has Website": "Yes",
        "Website Accessible": "Yes" if contacts["accessible"] else "No",
        "Has Email": "Yes" if contacts["emails"] else "No",
        "Has Phone": "Yes" if contacts["phones"] else "No",
        "Keyword": keyword,
        "Source": "Google Search",
        "Rating": 0,
        "Reviews": 0,
        "Extracted At": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        # Internal scoring fields
        "email": contacts["emails"][0] if contacts["emails"] else "",
        "phone": contacts["phones"][0] if contacts["phones"] else "",
        "website": result["url"],
        "website_accessible": contacts["accessible"],
        "facebook_url": contacts["socials"].get("facebook", ""),
        "instagram_url": contacts["socials"].get("instagram", ""),
        "linkedin_url": contacts["socials"].get("linkedin", ""),
        "twitter_url": contacts["socials"].get("twitter", ""),
        "rating": 0,
        "reviews_count": 0,
    }

    score, quality = scorer.score(lead)
    lead["Lead Score"] = score
    lead["Lead Quality"] = quality
    return lead


def _build_lead_from_facebook(fb_data: Dict, keyword: str,
                              city: str, country: str, scorer: LeadScorer) -> Dict:
    """Build a lead dict from Facebook extraction."""
    lead = {
        "Business Name": fb_data.get("business_name", "Unknown"),
        "Phone Number": fb_data.get("phone", "N/A") or "N/A",
        "All Phones": fb_data.get("phone", "N/A") or "N/A",
        "Email": fb_data.get("email", "N/A") or "N/A",
        "All Emails": fb_data.get("email", "N/A") or "N/A",
        "Website": fb_data.get("website", ""),
        "Domain": urllib.parse.urlparse(fb_data.get("website", "")).netloc if fb_data.get("website") else "",
        "Address": fb_data.get("address", f"{city}, {country}") or f"{city}, {country}",
        "Category": fb_data.get("category", "") or _guess_category(fb_data.get("business_name", ""), ""),
        "Snippet": "",
        "Facebook URL": fb_data.get("facebook_url", ""),
        "Instagram URL": "",
        "LinkedIn URL": "",
        "Twitter URL": "",
        "Has Website": "Yes" if fb_data.get("website") else "No",
        "Website Accessible": "Yes" if fb_data.get("accessible") else "No",
        "Has Email": "Yes" if fb_data.get("email") else "No",
        "Has Phone": "Yes" if fb_data.get("phone") else "No",
        "Keyword": keyword,
        "Source": "Facebook",
        "Rating": fb_data.get("rating", 0) or 0,
        "Reviews": fb_data.get("reviews_count", 0) or 0,
        # Internal scoring fields
        "email": fb_data.get("email", ""),
        "phone": fb_data.get("phone", ""),
        "website": fb_data.get("website", ""),
        "website_accessible": fb_data.get("accessible", False),
        "facebook_url": fb_data.get("facebook_url", ""),
        "instagram_url": "",
        "linkedin_url": "",
        "twitter_url": "",
        "rating": fb_data.get("rating", 0) or 0,
        "reviews_count": fb_data.get("reviews_count", 0) or 0,
    }

    score, quality = scorer.score(lead)
    lead["Lead Score"] = score
    lead["Lead Quality"] = quality
    return lead


def _build_lead_from_linkedin(li_data: Dict, keyword: str,
                              city: str, country: str, scorer: LeadScorer) -> Dict:
    """Build a lead dict from LinkedIn extraction."""
    lead = {
        "Business Name": li_data.get("business_name", "Unknown"),
        "Phone Number": li_data.get("phone", "N/A") or "N/A",
        "All Phones": li_data.get("phone", "N/A") or "N/A",
        "Email": li_data.get("email", "N/A") or "N/A",
        "All Emails": li_data.get("email", "N/A") or "N/A",
        "Website": li_data.get("website", ""),
        "Domain": urllib.parse.urlparse(li_data.get("website", "")).netloc if li_data.get("website") else "",
        "Address": li_data.get("address", f"{city}, {country}") or f"{city}, {country}",
        "Category": li_data.get("category", "") or _guess_category(li_data.get("business_name", ""), ""),
        "Snippet": "",
        "Facebook URL": "",
        "Instagram URL": "",
        "LinkedIn URL": li_data.get("linkedin_url", ""),
        "Twitter URL": "",
        "Has Website": "Yes" if li_data.get("website") else "No",
        "Website Accessible": "Yes" if li_data.get("accessible") else "No",
        "Has Email": "Yes" if li_data.get("email") else "No",
        "Has Phone": "Yes" if li_data.get("phone") else "No",
        "Keyword": keyword,
        "Source": "LinkedIn",
        "Rating": 0,
        "Reviews": 0,
        "Employee Count": li_data.get("employee_count", ""),
        # Internal scoring fields
        "email": li_data.get("email", ""),
        "phone": li_data.get("phone", ""),
        "website": li_data.get("website", ""),
        "website_accessible": li_data.get("accessible", False),
        "facebook_url": "",
        "instagram_url": "",
        "linkedin_url": li_data.get("linkedin_url", ""),
        "twitter_url": "",
        "rating": 0,
        "reviews_count": 0,
    }

    score, quality = scorer.score(lead)
    lead["Lead Score"] = score
    lead["Lead Quality"] = quality
    return lead


def _build_lead_from_places(places_data: Dict, keyword: str,
                            city: str, country: str,
                            extractor: ContactExtractor,
                            scorer: LeadScorer) -> Dict:
    """Build a lead from Google Places API data, enriched with website contact extraction."""
    website = places_data.get("website", "")
    contacts = {"emails": [], "phones": [], "socials": {}, "accessible": False}

    # If the place has a website, visit it to get email/socials
    if website:
        contacts = extractor.fetch_and_extract(website)

    phone = places_data.get("phone", "") or (contacts["phones"][0] if contacts["phones"] else "")
    email = contacts["emails"][0] if contacts["emails"] else ""

    lead = {
        "Business Name": places_data.get("business_name", "Unknown"),
        "Phone Number": phone or "N/A",
        "All Phones": phone or "N/A",
        "Email": email or "N/A",
        "All Emails": ", ".join(contacts["emails"]) if contacts["emails"] else "N/A",
        "Website": website,
        "Domain": urllib.parse.urlparse(website).netloc if website else "",
        "Address": places_data.get("address", f"{city}, {country}"),
        "Category": places_data.get("category", "Business"),
        "Snippet": "",
        "Facebook URL": contacts["socials"].get("facebook", ""),
        "Instagram URL": contacts["socials"].get("instagram", ""),
        "LinkedIn URL": contacts["socials"].get("linkedin", ""),
        "Twitter URL": contacts["socials"].get("twitter", ""),
        "Has Website": "Yes" if website else "No",
        "Website Accessible": "Yes" if contacts["accessible"] else ("N/A" if not website else "No"),
        "Has Email": "Yes" if email else "No",
        "Has Phone": "Yes" if phone else "No",
        "Keyword": keyword,
        "Source": "Google Places",
        "Rating": places_data.get("rating", 0) or 0,
        "Reviews": places_data.get("reviews_count", 0) or 0,
        "Google Maps URL": places_data.get("maps_url", ""),
        # Internal scoring fields
        "email": email,
        "phone": phone,
        "website": website,
        "website_accessible": contacts["accessible"],
        "facebook_url": contacts["socials"].get("facebook", ""),
        "instagram_url": contacts["socials"].get("instagram", ""),
        "linkedin_url": contacts["socials"].get("linkedin", ""),
        "twitter_url": contacts["socials"].get("twitter", ""),
        "rating": places_data.get("rating", 0) or 0,
        "reviews_count": places_data.get("reviews_count", 0) or 0,
    }

    score, quality = scorer.score(lead)
    lead["Lead Score"] = score
    lead["Lead Quality"] = quality
    return lead


def run_search(city: str, country: str, keywords_list: List[str], num_leads: int,
               use_places_api: bool = False, places_api_key: str = "") -> Dict:
    """
    Execute the full multi-source lead generation pipeline:
    1. Google Search for each keyword (websites + Facebook + LinkedIn)
    2. Google Places API for each keyword (optional)
    3. Visit each result URL to extract contacts
    4. Score and qualify leads
    """
    global all_leads

    google_scraper = GoogleSearchScraper()
    extractor = ContactExtractor()
    fb_extractor = FacebookExtractor()
    li_extractor = LinkedInExtractor()
    scorer = LeadScorer()

    places_scraper = None
    if use_places_api and places_api_key:
        places_scraper = GooglePlacesScraper(places_api_key)
        _add_log("Google Places API enabled", "success")

    all_results: List[Dict] = []
    seen_names: Set[str] = set()
    seen_domains: Set[str] = set()

    total_keywords = len(keywords_list)
    total_api_calls = 0

    for kw_idx, keyword in enumerate(keywords_list, 1):
        keyword = keyword.strip()
        if not keyword:
            continue

        if len(all_results) >= num_leads:
            break

        _add_log(f"[{kw_idx}/{total_keywords}] Searching for: \"{keyword}\" in {city}, {country}")

        # ---- GOOGLE SEARCH ----
        query = f"{keyword} in {city}, {country}"
        results_per_kw = max(10, num_leads // total_keywords + 5)

        search_results = google_scraper.search(query, num_results=min(results_per_kw, MAX_RESULTS_PER_KEYWORD))
        total_api_calls += 1

        websites = search_results["websites"]
        fb_pages = search_results["facebook"]
        li_pages = search_results["linkedin"]

        _add_log(f"  Google Search: {len(websites)} websites, {len(fb_pages)} Facebook, {len(li_pages)} LinkedIn", "success")

        # Process regular websites
        for r_idx, result in enumerate(websites):
            if len(all_results) >= num_leads:
                break
            domain = result["domain"]
            if domain in seen_domains:
                continue
            seen_domains.add(domain)

            _add_log(f"  [{r_idx+1}/{len(websites)}] Visiting {domain}...")
            contacts = extractor.fetch_and_extract(result["url"])
            lead = _build_lead_from_website(result, contacts, keyword, city, country, scorer)

            bname = lead["Business Name"].lower().strip()
            if bname not in seen_names:
                seen_names.add(bname)
                all_results.append(lead)
                if contacts["emails"]:
                    _add_log(f"    Email found: {contacts['emails'][0]}", "success")
                if contacts["phones"]:
                    _add_log(f"    Phone found: {contacts['phones'][0]}", "success")

        # Process Facebook pages
        for fb_result in fb_pages:
            if len(all_results) >= num_leads:
                break

            _add_log(f"  Extracting from Facebook: {fb_result['title'][:50]}...")
            fb_data = fb_extractor.extract(fb_result["url"], fb_result["title"], fb_result["snippet"])
            lead = _build_lead_from_facebook(fb_data, keyword, city, country, scorer)

            bname = lead["Business Name"].lower().strip()
            if bname and bname not in seen_names and bname != "unknown":
                seen_names.add(bname)
                all_results.append(lead)
                if fb_data.get("email"):
                    _add_log(f"    FB Email: {fb_data['email']}", "success")
                if fb_data.get("phone"):
                    _add_log(f"    FB Phone: {fb_data['phone']}", "success")
                if fb_data.get("website"):
                    _add_log(f"    FB Website: {fb_data['website']}", "success")

        # Process LinkedIn pages
        for li_result in li_pages:
            if len(all_results) >= num_leads:
                break

            _add_log(f"  Extracting from LinkedIn: {li_result['title'][:50]}...")
            li_data = li_extractor.extract(li_result["url"], li_result["title"], li_result["snippet"])
            lead = _build_lead_from_linkedin(li_data, keyword, city, country, scorer)

            bname = lead["Business Name"].lower().strip()
            if bname and bname not in seen_names and bname != "unknown":
                seen_names.add(bname)
                all_results.append(lead)
                if li_data.get("email"):
                    _add_log(f"    LI Email: {li_data['email']}", "success")
                if li_data.get("website"):
                    _add_log(f"    LI Website: {li_data['website']}", "success")

        # ---- GOOGLE PLACES API ----
        if places_scraper:
            _add_log(f"  Querying Google Places API for \"{keyword}\"...")
            places_query = f"{keyword} in {city}, {country}"
            places_per_kw = max(5, min(20, (num_leads - len(all_results)) // max(1, total_keywords - kw_idx + 1)))

            try:
                places_results = places_scraper.search(places_query, max_results=places_per_kw)
                total_api_calls += len(places_results) + 1  # 1 search + N detail calls

                _add_log(f"  Google Places: {len(places_results)} businesses found", "success")

                for p_idx, places_data in enumerate(places_results):
                    if len(all_results) >= num_leads:
                        break

                    bname = places_data["business_name"].lower().strip()
                    if bname in seen_names:
                        continue

                    _add_log(f"  [Places {p_idx+1}/{len(places_results)}] {places_data['business_name']}...")
                    lead = _build_lead_from_places(places_data, keyword, city, country, extractor, scorer)

                    seen_names.add(bname)
                    all_results.append(lead)

                    if lead.get("email") and lead["email"] != "N/A":
                        _add_log(f"    Places Email: {lead['email']}", "success")
                    if lead.get("phone") and lead["phone"] != "N/A":
                        _add_log(f"    Places Phone: {lead['Phone Number']}", "success")

            except Exception as e:
                _add_log(f"  Places API error: {str(e)[:100]}", "error")

        _add_log(f"  Keyword \"{keyword}\" complete -- Total leads: {len(all_results)}")

    # Sort by lead score descending
    all_results.sort(key=lambda x: x.get("Lead Score", 0), reverse=True)
    all_results = all_results[:num_leads]
    all_leads = all_results

    return {
        "leads": all_results,
        "total_leads": len(all_results),
        "api_calls": total_api_calls,
        "keywords_searched": total_keywords,
    }


# ============================================================
# WEB REQUEST ORCHESTRATION (called from blueprints/sales/routes.py)
# ============================================================
def execute_search(city: str, country: str, keywords: str, num_leads: int,
                    use_places: str = "false", api_key: str = "") -> Tuple[int, Dict]:
    """Run a search on behalf of an HTTP request. Returns (status_code, json_body)."""
    global search_logs

    search_logs = []

    if keywords.strip():
        keywords_list = [k.strip() for k in keywords.split(",") if k.strip()][:15]
    else:
        keywords_list = ["businesses", "services", "companies"]

    use_places_api = use_places.lower() in ("true", "1", "yes")

    _add_log("Starting lead generation search")
    _add_log(f"Location: {city}, {country}")
    _add_log(f"Keywords: {', '.join(keywords_list)}")
    _add_log(f"Target: {num_leads} leads")
    sources = ["Google Search", "Facebook Pages", "LinkedIn Pages"]
    if use_places_api and api_key:
        sources.append("Google Places API")
    _add_log(f"Sources: {', '.join(sources)}")

    search_id = hashlib.md5(f"{city}{country}{datetime.now()}".encode()).hexdigest()[:8]

    try:
        result = run_search(
            city, country, keywords_list, num_leads,
            use_places_api=use_places_api, places_api_key=api_key,
        )

        search_history.insert(0, {
            "search_id": search_id,
            "city": city,
            "country": country,
            "keywords": keywords if keywords else "All categories",
            "keywords_count": len(keywords_list),
            "leads_found": result["total_leads"],
            "api_calls": result["api_calls"],
            "sources": sources,
            "timestamp": datetime.now().isoformat(),
        })
        if len(search_history) > 20:
            search_history.pop()

        _add_log(f"Search complete! Found {result['total_leads']} leads", "success")

        return 200, {
            "success": True,
            "search_id": search_id,
            "total_leads": result["total_leads"],
            "api_calls": result["api_calls"],
            "leads": result["leads"],
            "message": f"Found {result['total_leads']} businesses in {city}, {country}",
        }

    except Exception as e:
        _add_log(f"Search failed: {str(e)}", "error")
        return 500, {"error": str(e), "success": False}


def get_stats() -> Dict:
    if not all_leads:
        return {"total": 0, "hot": 0, "warm": 0, "cold": 0,
                "has_email": 0, "has_phone": 0, "has_website": 0,
                "categories": {}, "avg_score": 0, "sources": {}}

    hot = len([l for l in all_leads if l.get("Lead Quality") == "Hot"])
    warm = len([l for l in all_leads if l.get("Lead Quality") == "Warm"])
    cold = len([l for l in all_leads if l.get("Lead Quality") == "Cold"])
    has_email = len([l for l in all_leads if l.get("Has Email") == "Yes"])
    has_phone = len([l for l in all_leads if l.get("Has Phone") == "Yes"])
    has_website = len([l for l in all_leads if l.get("Has Website") == "Yes"])

    categories = defaultdict(int)
    sources = defaultdict(int)
    for lead in all_leads:
        categories[lead.get("Category", "Unknown")] += 1
        sources[lead.get("Source", "Unknown")] += 1

    avg_score = sum(l.get("Lead Score", 0) for l in all_leads) // len(all_leads) if all_leads else 0

    return {
        "total": len(all_leads), "hot": hot, "warm": warm, "cold": cold,
        "has_email": has_email, "has_phone": has_phone, "has_website": has_website,
        "categories": dict(categories), "sources": dict(sources), "avg_score": avg_score,
    }


EXPORT_FIELDS = [
    "Business Name", "Email", "All Emails", "Phone Number", "All Phones",
    "Website", "Domain", "Address", "Category", "Lead Score", "Lead Quality",
    "Facebook URL", "Instagram URL", "LinkedIn URL", "Twitter URL",
    "Has Email", "Has Phone", "Website Accessible", "Source", "Rating", "Reviews",
    "Snippet", "Keyword", "Extracted At",
]


def export_csv_content() -> Optional[str]:
    if not all_leads:
        return None
    import io
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=EXPORT_FIELDS, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(all_leads)
    return buf.getvalue()


def export_excel_content() -> Optional[bytes]:
    if not all_leads:
        return None
    import io
    import pandas as pd
    from openpyxl.styles import PatternFill, Font, Alignment

    df = pd.DataFrame(all_leads)
    existing_fields = [f for f in EXPORT_FIELDS if f in df.columns]
    df = df[existing_fields]

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Leads", index=False)
        ws = writer.sheets["Leads"]

        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for col in range(1, len(existing_fields) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                try:
                    if cell.value and len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    return buf.getvalue()

