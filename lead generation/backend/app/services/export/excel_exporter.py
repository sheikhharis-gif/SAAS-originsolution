import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from datetime import datetime
import asyncio
from typing import List, Dict
import os

class ExcelExporter:
    """Professional Excel export with formatting"""
    
    def __init__(self):
        self.export_dir = "exports"
        os.makedirs(self.export_dir, exist_ok=True)
    
    async def export(self, leads: List[Dict], format: str = "excel") -> str:
        """Export leads to formatted Excel file"""
        
        # Convert leads to DataFrame
        df = await self._prepare_dataframe(leads)
        
        # Remove duplicates based on business name and phone
        df = df.drop_duplicates(subset=['business_name', 'phone_number'])
        
        # Create filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{self.export_dir}/leads_export_{timestamp}.xlsx"
        
        # Create Excel writer with xlsxwriter engine
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Leads', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Leads']
            
            # Apply formatting
            await self._apply_formatting(workbook, worksheet, df)
        
        return filename
    
    async def _prepare_dataframe(self, leads: List[Dict]) -> pd.DataFrame:
        """Prepare DataFrame with proper column ordering"""
        
        columns = [
            'business_name', 'owner_name', 'phone_number', 'whatsapp_number',
            'email', 'website', 'address', 'postal_code', 'category',
            'lead_score', 'lead_quality', 'rating', 'reviews_count',
            'facebook_url', 'instagram_url', 'linkedin_url', 'google_maps_url',
            'has_website', 'needs_website', 'needs_seo', 'needs_digital_marketing',
            'needs_crm', 'needs_tracking', 'needs_automation', 'status', 'tags'
        ]
        
        # Create DataFrame
        df = pd.DataFrame(leads)
        
        # Ensure all columns exist
        for col in columns:
            if col not in df.columns:
                df[col] = None
        
        # Reorder columns
        df = df[columns]
        
        # Clean data
        df = df.replace({float('nan'): None})
        
        return df
    
    async def _apply_formatting(self, workbook, worksheet, df):
        """Apply professional formatting to Excel sheet"""
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Format headers
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Auto-size columns
        for col in worksheet.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add conditional formatting for lead score
        score_column = df.columns.get_loc('lead_score') + 1
        score_range = f"{chr(64 + score_column)}2:{chr(64 + score_column)}{len(df) + 1}"
        
        # Color scale for lead scores
        color_scale = ColorScaleRule(
            start_type='min', start_color='FF9999',
            mid_type='percentile', mid_value=50, mid_color='FFFF99',
            end_type='max', end_color='99FF99'
        )
        worksheet.conditional_formatting.add(score_range, color_scale)
        
        # Add filters
        worksheet.auto_filter.ref = worksheet.dimensions
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
        
        # Add summary row with timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        summary_row = len(df) + 3
        worksheet[f'A{summary_row}'] = f"Export Generated: {timestamp}"
        worksheet[f'A{summary_row}'].font = Font(italic=True)
        worksheet[f'B{summary_row}'] = f"Total Leads: {len(df)}"
        worksheet[f'B{summary_row}'].font = Font(italic=True)