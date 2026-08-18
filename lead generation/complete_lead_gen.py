"""
LEAD GENERATION PLATFORM v4.0
Google Search + Google Places API + Facebook/LinkedIn Extraction
Multi-source, multi-keyword lead discovery with contact extraction
Run: python complete_lead_gen.py
Dashboard: http://localhost:8000
"""

from fastapi import FastAPI, Query, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, Response, JSONResponse
import uvicorn
import requests
from bs4 import BeautifulSoup
import csv
import json
import os
import re
import time
import random
import hashlib
import asyncio
import aiohttp
import urllib.parse
from datetime import datetime
from typing import List, Dict, Optional, Set, Tuple
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor

app = FastAPI(title="Lead Generation Platform", version="4.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============================================================
# CONFIGURATION
# ============================================================
SEARCH_DELAY_MIN = 1.5
SEARCH_DELAY_MAX = 3.0
WEBSITE_TIMEOUT = 4
MAX_CONCURRENT_VISITS = 5
MAX_RESULTS_PER_KEYWORD = 50
CONTACT_PAGES = ["/contact", "/about"]

# ============================================================
# DATA STORAGE
# ============================================================
all_leads: List[Dict] = []
search_history: List[Dict] = []
search_logs: List[Dict] = []

# ============================================================
# USER AGENT POOL
# ============================================================
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:126.0) Gecko/20100101 Firefox/126.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14.5; rv:126.0) Gecko/20100101 Firefox/126.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_5) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
]

# ============================================================
# BUSINESS CATEGORIES
# ============================================================
BUSINESS_CATEGORIES = {
    "Food & Dining": ["restaurants", "pizza places", "coffee shops", "bakery", "ice cream shops", "bars", "breweries"],
    "Beauty & Wellness": ["hair salons", "barbershops", "nail salons", "spas", "massage therapy"],
    "Health & Medical": ["dentists", "dental clinics", "doctors", "pediatricians", "dermatologists", "optometrists"],
    "Professional Services": ["lawyers", "accountants", "real estate agents", "financial advisors", "insurance agents"],
    "Home Services": ["plumbers", "electricians", "contractors", "handyman", "roofers", "painters", "landscapers"],
    "Automotive": ["auto repair", "mechanics", "car dealerships", "auto body shops", "car washes", "tire shops"],
    "Fitness": ["gyms", "fitness centers", "yoga studios", "personal trainers", "crossfit", "martial arts"],
    "Retail": ["clothing stores", "electronics stores", "furniture stores", "bookstores", "pharmacies"],
    "Education": ["tutoring", "music lessons", "dance classes", "art classes", "language schools"],
    "Pets": ["veterinarians", "pet grooming", "pet boarding", "pet stores"],
    "Events": ["event planners", "wedding planners", "photographers", "videographers", "caterers"],
    "Hotels & Travel": ["hotels", "motels", "bed and breakfasts", "vacation rentals"],
    "Cleaning": ["house cleaning", "carpet cleaning", "window cleaning", "commercial cleaning"],
}

# ============================================================
# LEAD SCORING RULES
# ============================================================
LEAD_SCORING_RULES = {
    "has_email": 15,
    "has_phone": 12,
    "has_website": 10,
    "has_social_media": 8,
    "has_multiple_contacts": 10,
    "rating_high": 15,
    "many_reviews": 10,
    "complete_info": 10,
    "category_bonus": 10,
}


# ============================================================
# HELPER UTILITIES
# ============================================================
def _add_log(message: str, level: str = "info"):
    """Append to the global log list."""
    search_logs.append({
        "timestamp": datetime.now().strftime("%H:%M:%S"),
        "message": message,
        "level": level,
    })
    icon = {"info": "[INFO]", "success": "[OK]", "warning": "[WARN]", "error": "[ERR]"}.get(level, "[LOG]")
    try:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {icon} {message}")
    except UnicodeEncodeError:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {icon} {message.encode('ascii', 'replace').decode()}")


def _extract_business_name(title: str, domain: str) -> str:
    """Try to derive a clean business name from page title."""
    if not title:
        return domain.split(".")[0].title()
    for sep in [" | ", " - ", " — ", " – ", " :: ", " » "]:
        if sep in title:
            title = title.split(sep)[0]
    return title.strip()


def _guess_category(title: str, snippet: str) -> str:
    """Guess business category from title and snippet text."""
    combined = (title + " " + snippet).lower()
    category_keywords = {
        "Restaurant": ["restaurant", "cafe", "bakery", "pizza", "sushi", "coffee", "dining", "food"],
        "Salon": ["salon", "barber", "hair", "nails", "beauty"],
        "Spa": ["spa", "massage", "wellness"],
        "Dental": ["dentist", "dental", "orthodont"],
        "Medical": ["doctor", "medical", "clinic", "physician", "health"],
        "Legal": ["lawyer", "attorney", "law firm", "legal"],
        "Real Estate": ["real estate", "realtor", "realty", "property"],
        "Accounting": ["accountant", "accounting", "cpa", "tax"],
        "Plumbing": ["plumber", "plumbing", "drain"],
        "Electrical": ["electrician", "electrical", "wiring"],
        "Construction": ["contractor", "construction", "remodel", "renovation"],
        "Auto Repair": ["auto repair", "mechanic", "car repair", "auto body"],
        "Gym": ["gym", "fitness", "workout", "crossfit"],
        "Hotel": ["hotel", "motel", "inn", "lodge", "resort"],
        "Retail": ["store", "shop", "boutique", "retail"],
        "Cleaning": ["cleaning", "janitorial", "maid service"],
        "Photography": ["photographer", "photography", "photo studio"],
        "Education": ["tutor", "school", "academy", "lesson", "training"],
        "Veterinary": ["vet", "veterinary", "animal hospital", "pet"],
        "Insurance": ["insurance", "coverage", "policy"],
    }
    for cat, kws in category_keywords.items():
        for kw in kws:
            if kw in combined:
                return cat
    return "Business"


# ============================================================
# GOOGLE SEARCH SCRAPER
# ============================================================
class GoogleSearchScraper:
    """Multi-engine Web Search Scraper (DuckDuckGo HTML + Google Search fallback)."""

    def __init__(self):
        self._request_count = 0

    def _get_headers(self) -> Dict[str, str]:
        ua = random.choice(USER_AGENTS)
        return {
            "User-Agent": ua,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
        }

    def search(self, query: str, num_results: int = 30) -> Dict[str, List[Dict]]:
        """
        Search Web for organic results, Facebook pages, and LinkedIn profiles.
        """
        websites: List[Dict] = []
        facebook: List[Dict] = []
        linkedin: List[Dict] = []
        seen_urls: Set[str] = set()

        # 1. Fetch Organic Web Results via DuckDuckGo HTML API
        ddg_results = self._search_ddg(query)
        for r in ddg_results:
            if r["url"] in seen_urls:
                continue
            seen_urls.add(r["url"])
            if r["type"] == "facebook":
                facebook.append(r)
            elif r["type"] == "linkedin":
                linkedin.append(r)
            else:
                websites.append(r)

        # 2. Specifically query Facebook pages if few found
        if len(facebook) < 3:
            fb_results = self._search_ddg(f"site:facebook.com {query}")
            for r in fb_results:
                if r["url"] not in seen_urls and r["type"] == "facebook":
                    seen_urls.add(r["url"])
                    facebook.append(r)

        # 3. Specifically query LinkedIn company pages if few found
        if len(linkedin) < 3:
            li_results = self._search_ddg(f"site:linkedin.com/company {query}")
            for r in li_results:
                if r["url"] not in seen_urls and r["type"] == "linkedin":
                    seen_urls.add(r["url"])
                    linkedin.append(r)

        # 4. Fallback to Google Search if DDG returns insufficient results
        if len(websites) < 3:
            google_results = self._search_google(query)
            for r in google_results:
                if r["url"] not in seen_urls:
                    seen_urls.add(r["url"])
                    if r["type"] == "facebook":
                        facebook.append(r)
                    elif r["type"] == "linkedin":
                        linkedin.append(r)
                    else:
                        websites.append(r)

        return {
            "websites": websites[:num_results],
            "facebook": facebook[:15],
            "linkedin": linkedin[:15],
        }

    def _search_ddg(self, query: str) -> List[Dict]:
        """Query DuckDuckGo HTML API and extract results."""
        results: List[Dict] = []
        url = "https://html.duckduckgo.com/html/"
        data = {"q": query}
        skip_domains = [
            "duckduckgo.com", "google.com", "youtube.com", "wikipedia.org",
            "pinterest.com", "reddit.com", "yelp.com", "yellowpages.com",
            "bbb.org", "tripadvisor.com", "consumeraffairs.com", "angi.com",
            "expertise.com", "thumbtack.com", "trustpilot.com",
        ]

        try:
            resp = requests.post(url, data=data, headers=self._get_headers(), timeout=8)
            self._request_count += 1
            if resp.status_code == 200:
                soup = BeautifulSoup(resp.text, "html.parser")
                for div in soup.select("div.result"):
                    a_url = div.select_one("a.result__url")
                    a_title = div.select_one("a.result__a")
                    a_snippet = div.select_one("a.result__snippet")

                    if not a_url or not a_url.get("href"):
                        continue

                    raw_href = a_url.get("href", "").strip()
                    m = re.search(r"uddg=([^&]+)", raw_href)
                    actual_url = urllib.parse.unquote(m.group(1)) if m else raw_href

                    if not actual_url.startswith("http"):
                        actual_url = "https://" + actual_url.lstrip("/")

                    domain = urllib.parse.urlparse(actual_url).netloc.lower().replace("www.", "")

                    if any(sd in domain for sd in skip_domains):
                        continue

                    title = a_title.get_text(strip=True) if a_title else ""
                    snippet = a_snippet.get_text(strip=True) if a_snippet else ""

                    res_type = "website"
                    if "facebook.com" in domain:
                        res_type = "facebook"
                    elif "linkedin.com" in domain:
                        res_type = "linkedin"

                    results.append({
                        "title": title,
                        "url": actual_url,
                        "snippet": snippet,
                        "domain": domain,
                        "type": res_type,
                    })

        except Exception as e:
            _add_log(f"Search provider note: {str(e)[:60]}", "warning")

        return results

    def _search_google(self, query: str) -> List[Dict]:
        """Fallback Google Search HTML parser."""
        results: List[Dict] = []
        params = {"q": query, "num": "15", "hl": "en"}
        url = "https://www.google.com/search?" + urllib.parse.urlencode(params)
        skip_domains = [
            "google.com", "google.co", "youtube.com", "wikipedia.org",
            "pinterest.com", "reddit.com", "yelp.com", "yellowpages.com",
            "bbb.org", "tripadvisor.com",
        ]

        try:
            resp = requests.get(url, headers=self._get_headers(), timeout=8)
            if resp.status_code == 200:
                soup = BeautifulSoup(resp.text, "html.parser")
                for div in soup.select("div.g, div.MjjYud"):
                    link_el = div.select_one("a[href]")
                    if not link_el:
                        continue
                    href = link_el.get("href", "")
                    if not href.startswith("http"):
                        continue
                    domain = urllib.parse.urlparse(href).netloc.lower().replace("www.", "")
                    if any(sd in domain for sd in skip_domains):
                        continue
                    title_el = div.select_one("h3")
                    title = title_el.get_text(strip=True) if title_el else ""
                    snippet_el = div.select_one("div.VwiC3b, span.aCOpRe")
                    snippet = snippet_el.get_text(strip=True) if snippet_el else ""

                    res_type = "website"
                    if "facebook.com" in domain:
                        res_type = "facebook"
                    elif "linkedin.com" in domain:
                        res_type = "linkedin"

                    results.append({
                        "title": title,
                        "url": href,
                        "snippet": snippet,
                        "domain": domain,
                        "type": res_type,
                    })
        except Exception:
            pass

        return results


# ============================================================
# GOOGLE PLACES API SCRAPER
# ============================================================
class GooglePlacesScraper:
    """Uses Google Places Text Search + Details API for structured business data."""

    def __init__(self, api_key: str):
        self.api_key = api_key

    def search(self, query: str, max_results: int = 20) -> List[Dict]:
        """Search Google Places and return enriched business data."""
        results: List[Dict] = []
        next_page_token = None

        url = "https://maps.googleapis.com/maps/api/place/textsearch/json"
        params = {"query": query, "key": self.api_key}

        # Fetch up to 3 pages (60 results max from Places API)
        for page_num in range(3):
            if len(results) >= max_results:
                break

            if page_num > 0:
                if not next_page_token:
                    break
                time.sleep(2)  # Required delay for pagetoken
                params["pagetoken"] = next_page_token

            try:
                resp = requests.get(url, params=params, timeout=15)
                if resp.status_code != 200:
                    _add_log(f"Places API error: status {resp.status_code}", "error")
                    break

                data = resp.json()
                status = data.get("status", "")

                if status == "REQUEST_DENIED":
                    _add_log("Places API: Invalid API key or API not enabled", "error")
                    break
                elif status == "OVER_QUERY_LIMIT":
                    _add_log("Places API: Query limit exceeded", "warning")
                    break
                elif status not in ("OK", "ZERO_RESULTS"):
                    _add_log(f"Places API status: {status}", "warning")
                    break

                places = data.get("results", [])
                next_page_token = data.get("next_page_token")

                for place in places:
                    if len(results) >= max_results:
                        break
                    details = self._get_details(place.get("place_id", ""))
                    results.append(self._build_lead(place, details))

            except Exception as e:
                _add_log(f"Places API error: {str(e)[:100]}", "error")
                break

        return results

    def _get_details(self, place_id: str) -> Dict:
        """Get detailed place information."""
        if not place_id:
            return {}

        url = "https://maps.googleapis.com/maps/api/place/details/json"
        params = {
            "place_id": place_id,
            "key": self.api_key,
            "fields": "formatted_phone_number,website,opening_hours,url",
        }

        try:
            resp = requests.get(url, params=params, timeout=10)
            if resp.status_code == 200:
                result = resp.json().get("result", {})
                return {
                    "phone": result.get("formatted_phone_number", ""),
                    "website": result.get("website", ""),
                    "opening_hours": result.get("opening_hours", {}).get("weekday_text", []),
                    "maps_url": result.get("url", ""),
                }
        except Exception:
            pass

        return {}

    def _extract_postal_code(self, address: str) -> str:
        """Extract postal code from address string."""
        match = re.search(r"\b\d{5}(?:-\d{4})?\b", address)
        return match.group() if match else ""

    def _build_lead(self, place: Dict, details: Dict) -> Dict:
        """Build a lead dict from Places API data."""
        address = place.get("formatted_address", "")
        rating = place.get("rating", 0)
        reviews = place.get("user_ratings_total", 0)
        website = details.get("website", "")
        phone = details.get("phone", "")
        types = place.get("types", [])

        # Determine category
        type_map = {
            "restaurant": "Restaurant", "cafe": "Cafe", "bar": "Bar",
            "hair_care": "Salon", "spa": "Spa", "gym": "Gym",
            "dentist": "Dental", "doctor": "Medical", "lawyer": "Legal",
            "real_estate_agency": "Real Estate", "accounting": "Accounting",
            "plumber": "Plumbing", "electrician": "Electrical",
            "car_repair": "Auto Repair", "hotel": "Hotel", "store": "Retail",
        }
        category = "Business"
        for t in types:
            if t in type_map:
                category = type_map[t]
                break

        return {
            "business_name": place.get("name", "Unknown"),
            "phone": phone,
            "website": website,
            "address": address,
            "postal_code": self._extract_postal_code(address),
            "rating": rating,
            "reviews_count": reviews,
            "category": category,
            "maps_url": details.get("maps_url", f"https://maps.google.com/?q={place.get('place_id', '')}"),
            "lat": place.get("geometry", {}).get("location", {}).get("lat", 0),
            "lng": place.get("geometry", {}).get("location", {}).get("lng", 0),
            "source": "google_places",
        }


# ============================================================
# FACEBOOK PAGE EXTRACTOR
# ============================================================
class FacebookExtractor:
    """Extract business info from Facebook pages found in search results."""

    def extract(self, url: str, title: str = "", snippet: str = "") -> Dict:
        """
        Extract business info from a Facebook page.
        Uses page HTML + structured data extraction.
        """
        result = {
            "business_name": "",
            "phone": "",
            "email": "",
            "website": "",
            "address": "",
            "category": "",
            "facebook_url": url,
            "source": "facebook",
            "accessible": False,
        }

        # Extract business name from title
        fb_name = title
        for suffix in [" - Facebook", " | Facebook", "- Home | Facebook",
                        " - Home", " - About", " - Posts", " - Reviews"]:
            fb_name = fb_name.replace(suffix, "")
        result["business_name"] = fb_name.strip()

        # Try to extract info from snippet
        snippet_info = self._parse_snippet(snippet)
        result.update({k: v for k, v in snippet_info.items() if v})

        # Try to fetch the page
        try:
            headers = {
                "User-Agent": random.choice(USER_AGENTS),
                "Accept": "text/html,application/xhtml+xml",
                "Accept-Language": "en-US,en;q=0.9",
            }
            resp = requests.get(url, headers=headers, timeout=WEBSITE_TIMEOUT,
                              allow_redirects=True, verify=False)

            if resp.status_code == 200 and "text/html" in resp.headers.get("content-type", ""):
                result["accessible"] = True
                page_info = self._parse_page(resp.text)
                # Merge — page data overrides snippet data if non-empty
                for k, v in page_info.items():
                    if v:
                        result[k] = v

        except Exception:
            pass

        return result

    def _parse_snippet(self, snippet: str) -> Dict:
        """Extract phone/email/address from Google snippet of FB page."""
        info: Dict = {}

        # Phone patterns in snippets
        phone_match = re.search(
            r"(?:Phone|Tel|Call)[:\s]*([+\d\s\-().]{7,20})", snippet, re.IGNORECASE
        )
        if phone_match:
            info["phone"] = phone_match.group(1).strip()

        # General phone pattern
        if not info.get("phone"):
            phone_match2 = re.search(
                r"(?:\+?\d{1,3}[\s\-.]?)?\(?\d{2,4}\)?[\s\-.]?\d{3,4}[\s\-.]?\d{3,4}", snippet
            )
            if phone_match2:
                info["phone"] = phone_match2.group().strip()

        # Email
        email_match = re.search(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", snippet)
        if email_match:
            info["email"] = email_match.group()

        # Rating
        rating_match = re.search(r"(\d+\.?\d*)\s*(?:stars?|rating)", snippet, re.IGNORECASE)
        if rating_match:
            try:
                info["rating"] = float(rating_match.group(1))
            except ValueError:
                pass

        # Reviews
        reviews_match = re.search(r"(\d+[,\d]*)\s*(?:reviews?|ratings?)", snippet, re.IGNORECASE)
        if reviews_match:
            try:
                info["reviews_count"] = int(reviews_match.group(1).replace(",", ""))
            except ValueError:
                pass

        return info

    def _parse_page(self, html: str) -> Dict:
        """Extract structured data from Facebook page HTML."""
        info: Dict = {}
        soup = BeautifulSoup(html, "html.parser")

        # Phone numbers from page content
        phone_pattern = re.compile(
            r"(?:\+?\d{1,3}[\s\-.]?)?\(?\d{2,4}\)?[\s\-.]?\d{3,4}[\s\-.]?\d{3,4}"
        )
        email_pattern = re.compile(
            r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", re.IGNORECASE
        )
        website_pattern = re.compile(
            r'https?://(?!(?:www\.)?facebook\.com)[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}[/\w.-]*',
            re.IGNORECASE,
        )

        text = soup.get_text(" ", strip=True)

        # Emails (exclude facebook emails)
        emails = email_pattern.findall(text)
        emails = [e for e in emails if "facebook" not in e.lower() and "fb.com" not in e.lower()]
        if emails:
            info["email"] = emails[0]

        # Phone
        phones = phone_pattern.findall(text)
        phones = [p.strip() for p in phones if len(re.sub(r"\D", "", p)) >= 7]
        if phones:
            info["phone"] = phones[0]

        # External website
        for a_tag in soup.select("a[href]"):
            href = a_tag.get("href", "")
            if href.startswith("http") and "facebook.com" not in href and "fb.com" not in href:
                # Check if it looks like a real website
                parsed = urllib.parse.urlparse(href)
                if parsed.netloc and "." in parsed.netloc:
                    info["website"] = href
                    break

        # Try JSON-LD structured data
        for script in soup.select("script[type='application/ld+json']"):
            try:
                ld_data = json.loads(script.string or "")
                if isinstance(ld_data, dict):
                    if ld_data.get("telephone"):
                        info["phone"] = ld_data["telephone"]
                    if ld_data.get("email"):
                        info["email"] = ld_data["email"]
                    if ld_data.get("url") and "facebook.com" not in ld_data.get("url", ""):
                        info["website"] = ld_data["url"]
                    addr = ld_data.get("address", {})
                    if isinstance(addr, dict) and addr.get("streetAddress"):
                        info["address"] = f"{addr.get('streetAddress', '')}, {addr.get('addressLocality', '')}"
            except Exception:
                continue

        # Try meta tags
        for meta in soup.select("meta[property], meta[name]"):
            prop = meta.get("property", "") or meta.get("name", "")
            content = meta.get("content", "")
            if prop == "og:description" and content and not info.get("address"):
                # Often contains category and address info
                if any(kw in content.lower() for kw in ["located", "address", "street", "avenue", "road"]):
                    info["address"] = content[:200]

        return info


# ============================================================
# LINKEDIN COMPANY EXTRACTOR
# ============================================================
class LinkedInExtractor:
    """Extract company info from LinkedIn pages found in search results."""

    def extract(self, url: str, title: str = "", snippet: str = "") -> Dict:
        """Extract company info from a LinkedIn company page."""
        result = {
            "business_name": "",
            "phone": "",
            "email": "",
            "website": "",
            "address": "",
            "category": "",
            "linkedin_url": url,
            "employee_count": "",
            "source": "linkedin",
            "accessible": False,
        }

        # Extract company name from title
        li_name = title
        for suffix in [" | LinkedIn", " - LinkedIn", " on LinkedIn",
                        " - Company Profile", " Overview", " - Overview"]:
            li_name = li_name.replace(suffix, "")
        result["business_name"] = li_name.strip()

        # Parse snippet for info
        snippet_info = self._parse_snippet(snippet)
        result.update({k: v for k, v in snippet_info.items() if v})

        # Try to fetch the page
        try:
            headers = {
                "User-Agent": random.choice(USER_AGENTS),
                "Accept": "text/html,application/xhtml+xml",
                "Accept-Language": "en-US,en;q=0.9",
            }
            resp = requests.get(url, headers=headers, timeout=WEBSITE_TIMEOUT,
                              allow_redirects=True, verify=False)

            if resp.status_code == 200 and "text/html" in resp.headers.get("content-type", ""):
                result["accessible"] = True
                page_info = self._parse_page(resp.text)
                for k, v in page_info.items():
                    if v:
                        result[k] = v

        except Exception:
            pass

        return result

    def _parse_snippet(self, snippet: str) -> Dict:
        """Extract info from Google snippet of LinkedIn page."""
        info: Dict = {}

        # Employee count
        emp_match = re.search(r"(\d+[,\d]*)\s*(?:employees?|followers?|staff)", snippet, re.IGNORECASE)
        if emp_match:
            info["employee_count"] = emp_match.group(1)

        # Industry/category
        industry_match = re.search(
            r"(?:Industry|Sector)[:\s]*([A-Za-z\s&,]+?)(?:\.|;|\d|$)", snippet, re.IGNORECASE
        )
        if industry_match:
            info["category"] = industry_match.group(1).strip()

        # Location
        loc_match = re.search(
            r"(?:Location|Based in|Headquartered in|Located in)[:\s]*([^.;]+)", snippet, re.IGNORECASE
        )
        if loc_match:
            info["address"] = loc_match.group(1).strip()

        # Phone
        phone_match = re.search(
            r"(?:\+?\d{1,3}[\s\-.]?)?\(?\d{2,4}\)?[\s\-.]?\d{3,4}[\s\-.]?\d{3,4}", snippet
        )
        if phone_match:
            info["phone"] = phone_match.group().strip()

        # Website URL in snippet
        web_match = re.search(
            r"(https?://(?!linkedin\.com)[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}[/\w.-]*)", snippet
        )
        if web_match:
            info["website"] = web_match.group(1)

        return info

    def _parse_page(self, html: str) -> Dict:
        """Extract structured data from LinkedIn page HTML."""
        info: Dict = {}
        soup = BeautifulSoup(html, "html.parser")

        # Try JSON-LD
        for script in soup.select("script[type='application/ld+json']"):
            try:
                ld = json.loads(script.string or "")
                if isinstance(ld, dict):
                    if ld.get("@type") in ("Organization", "Corporation", "LocalBusiness"):
                        if ld.get("telephone"):
                            info["phone"] = ld["telephone"]
                        if ld.get("email"):
                            info["email"] = ld["email"]
                        if ld.get("url") and "linkedin.com" not in ld.get("url", ""):
                            info["website"] = ld["url"]
                        if ld.get("numberOfEmployees"):
                            emp = ld["numberOfEmployees"]
                            if isinstance(emp, dict):
                                info["employee_count"] = str(emp.get("value", ""))
                            else:
                                info["employee_count"] = str(emp)
                        addr = ld.get("address", {})
                        if isinstance(addr, dict) and addr.get("addressLocality"):
                            info["address"] = f"{addr.get('streetAddress', '')}, {addr.get('addressLocality', '')}, {addr.get('addressRegion', '')}".strip(", ")
            except Exception:
                continue

        # Try meta tags
        for meta in soup.select("meta[property], meta[name]"):
            prop = meta.get("property", "") or meta.get("name", "")
            content = meta.get("content", "")
            if not content:
                continue
            if prop == "og:title" and not info.get("business_name"):
                name = content
                for suffix in [" | LinkedIn", " - LinkedIn"]:
                    name = name.replace(suffix, "")
                info["business_name"] = name.strip()
            if prop == "og:description" and content:
                # Extract website from description
                web_match = re.search(
                    r"(https?://(?!linkedin\.com)[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})", content
                )
                if web_match and not info.get("website"):
                    info["website"] = web_match.group(1)

        # Look for external links
        email_pattern = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", re.IGNORECASE)
        text = soup.get_text(" ", strip=True)
        emails = email_pattern.findall(text)
        emails = [e for e in emails if "linkedin" not in e.lower()]
        if emails and not info.get("email"):
            info["email"] = emails[0]

        return info


# ============================================================
# CONTACT EXTRACTOR (for regular websites)
# ============================================================
class ContactExtractor:
    """Visits websites to extract emails, phone numbers, and social links."""

    EMAIL_PATTERN = re.compile(
        r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", re.IGNORECASE
    )
    PHONE_PATTERN = re.compile(
        r"(?:(?:\+?\d{1,3}[\s\-.]?)?\(?\d{2,4}\)?[\s\-.]?\d{3,4}[\s\-.]?\d{3,4})"
    )
    SOCIAL_PATTERNS = {
        "facebook": re.compile(r"https?://(?:www\.)?facebook\.com/[a-zA-Z0-9._\-/]+", re.IGNORECASE),
        "instagram": re.compile(r"https?://(?:www\.)?instagram\.com/[a-zA-Z0-9._\-/]+", re.IGNORECASE),
        "linkedin": re.compile(r"https?://(?:www\.)?linkedin\.com/(?:in|company)/[a-zA-Z0-9._\-/]+", re.IGNORECASE),
        "twitter": re.compile(r"https?://(?:www\.)?(?:twitter\.com|x\.com)/[a-zA-Z0-9._\-/]+", re.IGNORECASE),
    }
    JUNK_EMAILS = re.compile(
        r"(?:example\.com|test\.com|email\.com|sentry\.io|wixpress|"
        r"schema\.org|googleapis|w3\.org|placeholder|noreply|no-reply|"
        r"\.png|\.jpg|\.gif|\.svg|\.webp|\.css|\.js)",
        re.IGNORECASE,
    )

    def _clean_emails(self, raw_emails: List[str]) -> List[str]:
        clean: List[str] = []
        seen: Set[str] = set()
        for email in raw_emails:
            email = email.lower().strip().rstrip(".")
            if email in seen:
                continue
            if self.JUNK_EMAILS.search(email):
                continue
            if len(email) < 6 or len(email) > 254:
                continue
            parts = email.rsplit(".", 1)
            if len(parts) != 2 or len(parts[1]) < 2:
                continue
            seen.add(email)
            clean.append(email)
        return clean

    def _clean_phones(self, raw_phones: List[str]) -> List[str]:
        clean: List[str] = []
        seen: Set[str] = set()
        for phone in raw_phones:
            phone = re.sub(r"\s+", " ", phone).strip()
            digits_only = re.sub(r"\D", "", phone)
            if len(digits_only) < 7 or len(digits_only) > 15:
                continue
            if digits_only in seen:
                continue
            seen.add(digits_only)
            clean.append(phone)
        return clean

    def extract_from_html(self, html: str, base_url: str = "") -> Dict:
        emails = self._clean_emails(self.EMAIL_PATTERN.findall(html))
        phones = self._clean_phones(self.PHONE_PATTERN.findall(html))
        socials = {}
        for platform, pattern in self.SOCIAL_PATTERNS.items():
            matches = pattern.findall(html)
            if matches:
                socials[platform] = matches[0].rstrip("/")
        return {"emails": emails[:5], "phones": phones[:5], "socials": socials}

    def fetch_and_extract(self, url: str) -> Dict:
        result = {"emails": [], "phones": [], "socials": {}, "accessible": False}
        try:
            headers = {"User-Agent": random.choice(USER_AGENTS)}
            resp = requests.get(url, headers=headers, timeout=WEBSITE_TIMEOUT,
                              allow_redirects=True, verify=False)
            if resp.status_code == 200 and "text/html" in resp.headers.get("content-type", ""):
                result["accessible"] = True
                extracted = self.extract_from_html(resp.text, url)
                result["emails"] = extracted["emails"]
                result["phones"] = extracted["phones"]
                result["socials"] = extracted["socials"]

                if not result["emails"]:
                    parsed = urllib.parse.urlparse(url)
                    base = f"{parsed.scheme}://{parsed.netloc}"
                    for contact_path in CONTACT_PAGES:
                        try:
                            r2 = requests.get(base + contact_path, headers=headers,
                                            timeout=2.5, allow_redirects=True, verify=False)
                            if r2.status_code == 200 and "text/html" in r2.headers.get("content-type", ""):
                                extra = self.extract_from_html(r2.text, base + contact_path)
                                result["emails"].extend(extra["emails"])
                                result["phones"].extend(extra["phones"])
                                for k, v in extra["socials"].items():
                                    if k not in result["socials"]:
                                        result["socials"][k] = v
                                if result["emails"]:
                                    break
                        except Exception:
                            continue

                result["emails"] = list(dict.fromkeys(result["emails"]))[:5]
                result["phones"] = list(dict.fromkeys(result["phones"]))[:5]

        except Exception:
            pass
        return result


# ============================================================
# LEAD SCORER
# ============================================================
class LeadScorer:
    """Score and qualify discovered leads."""

    @staticmethod
    def score(lead: Dict) -> Tuple[int, str]:
        score = 0

        if lead.get("email"):
            score += LEAD_SCORING_RULES["has_email"]
        if lead.get("phone"):
            score += LEAD_SCORING_RULES["has_phone"]
        if lead.get("website") and lead.get("website_accessible", True):
            score += LEAD_SCORING_RULES["has_website"]

        social_count = sum(
            1 for k in ["facebook_url", "instagram_url", "linkedin_url", "twitter_url"]
            if lead.get(k)
        )
        if social_count > 0:
            score += min(LEAD_SCORING_RULES["has_social_media"], social_count * 3)

        contact_channels = sum([
            bool(lead.get("email")),
            bool(lead.get("phone")),
            social_count > 0,
            bool(lead.get("website")),
        ])
        if contact_channels >= 3:
            score += LEAD_SCORING_RULES["has_multiple_contacts"]

        rating = lead.get("rating", 0) or 0
        if rating >= 4.0:
            score += LEAD_SCORING_RULES["rating_high"]
        elif rating >= 3.0:
            score += 8

        reviews = lead.get("reviews_count", 0) or 0
        if reviews >= 50:
            score += LEAD_SCORING_RULES["many_reviews"]
        elif reviews >= 20:
            score += 5

        score = min(100, max(0, score))
        quality = "Hot" if score >= 70 else ("Warm" if score >= 40 else "Cold")
        return score, quality


# ============================================================
# SEARCH ORCHESTRATION
# ============================================================
def _build_lead_from_website(result: Dict, contacts: Dict, keyword: str,
                             city: str, country: str, scorer: LeadScorer) -> Dict:
    """Build a lead dict from a website result + extracted contacts."""
    business_name = _extract_business_name(result["title"], result["domain"])
    category = _guess_category(result["title"], result["snippet"])

    lead = {
        "Business Name": business_name,
        "Phone Number": contacts["phones"][0] if contacts["phones"] else "N/A",
        "All Phones": ", ".join(contacts["phones"]) if contacts["phones"] else "N/A",
        "Email": contacts["emails"][0] if contacts["emails"] else "N/A",
        "All Emails": ", ".join(contacts["emails"]) if contacts["emails"] else "N/A",
        "Website": result["url"],
        "Domain": result["domain"],
        "Address": f"{city}, {country}",
        "Category": category,
        "Snippet": result["snippet"][:200] if result.get("snippet") else "",
        "Facebook URL": contacts["socials"].get("facebook", ""),
        "Instagram URL": contacts["socials"].get("instagram", ""),
        "LinkedIn URL": contacts["socials"].get("linkedin", ""),
        "Twitter URL": contacts["socials"].get("twitter", ""),
        "Has Website": "Yes",
        "Website Accessible": "Yes" if contacts["accessible"] else "No",
        "Has Email": "Yes" if contacts["emails"] else "No",
        "Has Phone": "Yes" if contacts["phones"] else "No",
        "Keyword": keyword,
        "Source": "Google Search",
        "Rating": 0,
        "Reviews": 0,
        "Extracted At": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        # Internal scoring fields
        "email": contacts["emails"][0] if contacts["emails"] else "",
        "phone": contacts["phones"][0] if contacts["phones"] else "",
        "website": result["url"],
        "website_accessible": contacts["accessible"],
        "facebook_url": contacts["socials"].get("facebook", ""),
        "instagram_url": contacts["socials"].get("instagram", ""),
        "linkedin_url": contacts["socials"].get("linkedin", ""),
        "twitter_url": contacts["socials"].get("twitter", ""),
        "rating": 0,
        "reviews_count": 0,
    }

    score, quality = scorer.score(lead)
    lead["Lead Score"] = score
    lead["Lead Quality"] = quality
    return lead


def _build_lead_from_facebook(fb_data: Dict, keyword: str,
                              city: str, country: str, scorer: LeadScorer) -> Dict:
    """Build a lead dict from Facebook extraction."""
    lead = {
        "Business Name": fb_data.get("business_name", "Unknown"),
        "Phone Number": fb_data.get("phone", "N/A") or "N/A",
        "All Phones": fb_data.get("phone", "N/A") or "N/A",
        "Email": fb_data.get("email", "N/A") or "N/A",
        "All Emails": fb_data.get("email", "N/A") or "N/A",
        "Website": fb_data.get("website", ""),
        "Domain": urllib.parse.urlparse(fb_data.get("website", "")).netloc if fb_data.get("website") else "",
        "Address": fb_data.get("address", f"{city}, {country}") or f"{city}, {country}",
        "Category": fb_data.get("category", "") or _guess_category(fb_data.get("business_name", ""), ""),
        "Snippet": "",
        "Facebook URL": fb_data.get("facebook_url", ""),
        "Instagram URL": "",
        "LinkedIn URL": "",
        "Twitter URL": "",
        "Has Website": "Yes" if fb_data.get("website") else "No",
        "Website Accessible": "Yes" if fb_data.get("accessible") else "No",
        "Has Email": "Yes" if fb_data.get("email") else "No",
        "Has Phone": "Yes" if fb_data.get("phone") else "No",
        "Keyword": keyword,
        "Source": "Facebook",
        "Rating": fb_data.get("rating", 0) or 0,
        "Reviews": fb_data.get("reviews_count", 0) or 0,
        # Internal scoring fields
        "email": fb_data.get("email", ""),
        "phone": fb_data.get("phone", ""),
        "website": fb_data.get("website", ""),
        "website_accessible": fb_data.get("accessible", False),
        "facebook_url": fb_data.get("facebook_url", ""),
        "instagram_url": "",
        "linkedin_url": "",
        "twitter_url": "",
        "rating": fb_data.get("rating", 0) or 0,
        "reviews_count": fb_data.get("reviews_count", 0) or 0,
    }

    score, quality = scorer.score(lead)
    lead["Lead Score"] = score
    lead["Lead Quality"] = quality
    return lead


def _build_lead_from_linkedin(li_data: Dict, keyword: str,
                              city: str, country: str, scorer: LeadScorer) -> Dict:
    """Build a lead dict from LinkedIn extraction."""
    lead = {
        "Business Name": li_data.get("business_name", "Unknown"),
        "Phone Number": li_data.get("phone", "N/A") or "N/A",
        "All Phones": li_data.get("phone", "N/A") or "N/A",
        "Email": li_data.get("email", "N/A") or "N/A",
        "All Emails": li_data.get("email", "N/A") or "N/A",
        "Website": li_data.get("website", ""),
        "Domain": urllib.parse.urlparse(li_data.get("website", "")).netloc if li_data.get("website") else "",
        "Address": li_data.get("address", f"{city}, {country}") or f"{city}, {country}",
        "Category": li_data.get("category", "") or _guess_category(li_data.get("business_name", ""), ""),
        "Snippet": "",
        "Facebook URL": "",
        "Instagram URL": "",
        "LinkedIn URL": li_data.get("linkedin_url", ""),
        "Twitter URL": "",
        "Has Website": "Yes" if li_data.get("website") else "No",
        "Website Accessible": "Yes" if li_data.get("accessible") else "No",
        "Has Email": "Yes" if li_data.get("email") else "No",
        "Has Phone": "Yes" if li_data.get("phone") else "No",
        "Keyword": keyword,
        "Source": "LinkedIn",
        "Rating": 0,
        "Reviews": 0,
        "Employee Count": li_data.get("employee_count", ""),
        # Internal scoring fields
        "email": li_data.get("email", ""),
        "phone": li_data.get("phone", ""),
        "website": li_data.get("website", ""),
        "website_accessible": li_data.get("accessible", False),
        "facebook_url": "",
        "instagram_url": "",
        "linkedin_url": li_data.get("linkedin_url", ""),
        "twitter_url": "",
        "rating": 0,
        "reviews_count": 0,
    }

    score, quality = scorer.score(lead)
    lead["Lead Score"] = score
    lead["Lead Quality"] = quality
    return lead


def _build_lead_from_places(places_data: Dict, keyword: str,
                            city: str, country: str,
                            extractor: ContactExtractor,
                            scorer: LeadScorer) -> Dict:
    """Build a lead from Google Places API data, enriched with website contact extraction."""
    website = places_data.get("website", "")
    contacts = {"emails": [], "phones": [], "socials": {}, "accessible": False}

    # If the place has a website, visit it to get email/socials
    if website:
        contacts = extractor.fetch_and_extract(website)

    phone = places_data.get("phone", "") or (contacts["phones"][0] if contacts["phones"] else "")
    email = contacts["emails"][0] if contacts["emails"] else ""

    lead = {
        "Business Name": places_data.get("business_name", "Unknown"),
        "Phone Number": phone or "N/A",
        "All Phones": phone or "N/A",
        "Email": email or "N/A",
        "All Emails": ", ".join(contacts["emails"]) if contacts["emails"] else "N/A",
        "Website": website,
        "Domain": urllib.parse.urlparse(website).netloc if website else "",
        "Address": places_data.get("address", f"{city}, {country}"),
        "Category": places_data.get("category", "Business"),
        "Snippet": "",
        "Facebook URL": contacts["socials"].get("facebook", ""),
        "Instagram URL": contacts["socials"].get("instagram", ""),
        "LinkedIn URL": contacts["socials"].get("linkedin", ""),
        "Twitter URL": contacts["socials"].get("twitter", ""),
        "Has Website": "Yes" if website else "No",
        "Website Accessible": "Yes" if contacts["accessible"] else ("N/A" if not website else "No"),
        "Has Email": "Yes" if email else "No",
        "Has Phone": "Yes" if phone else "No",
        "Keyword": keyword,
        "Source": "Google Places",
        "Rating": places_data.get("rating", 0) or 0,
        "Reviews": places_data.get("reviews_count", 0) or 0,
        "Google Maps URL": places_data.get("maps_url", ""),
        # Internal scoring fields
        "email": email,
        "phone": phone,
        "website": website,
        "website_accessible": contacts["accessible"],
        "facebook_url": contacts["socials"].get("facebook", ""),
        "instagram_url": contacts["socials"].get("instagram", ""),
        "linkedin_url": contacts["socials"].get("linkedin", ""),
        "twitter_url": contacts["socials"].get("twitter", ""),
        "rating": places_data.get("rating", 0) or 0,
        "reviews_count": places_data.get("reviews_count", 0) or 0,
    }

    score, quality = scorer.score(lead)
    lead["Lead Score"] = score
    lead["Lead Quality"] = quality
    return lead


def run_search(city: str, country: str, keywords_list: List[str], num_leads: int,
               use_places_api: bool = False, places_api_key: str = "") -> Dict:
    """
    Execute the full multi-source lead generation pipeline:
    1. Google Search for each keyword (websites + Facebook + LinkedIn)
    2. Google Places API for each keyword (optional)
    3. Visit each result URL to extract contacts
    4. Score and qualify leads
    """
    global all_leads

    google_scraper = GoogleSearchScraper()
    extractor = ContactExtractor()
    fb_extractor = FacebookExtractor()
    li_extractor = LinkedInExtractor()
    scorer = LeadScorer()

    places_scraper = None
    if use_places_api and places_api_key:
        places_scraper = GooglePlacesScraper(places_api_key)
        _add_log("Google Places API enabled", "success")

    all_results: List[Dict] = []
    seen_names: Set[str] = set()
    seen_domains: Set[str] = set()

    total_keywords = len(keywords_list)
    total_api_calls = 0

    for kw_idx, keyword in enumerate(keywords_list, 1):
        keyword = keyword.strip()
        if not keyword:
            continue

        if len(all_results) >= num_leads:
            break

        _add_log(f"[{kw_idx}/{total_keywords}] Searching for: \"{keyword}\" in {city}, {country}")

        # ---- GOOGLE SEARCH ----
        query = f"{keyword} in {city}, {country}"
        results_per_kw = max(10, num_leads // total_keywords + 5)

        search_results = google_scraper.search(query, num_results=min(results_per_kw, MAX_RESULTS_PER_KEYWORD))
        total_api_calls += 1

        websites = search_results["websites"]
        fb_pages = search_results["facebook"]
        li_pages = search_results["linkedin"]

        _add_log(f"  Google Search: {len(websites)} websites, {len(fb_pages)} Facebook, {len(li_pages)} LinkedIn", "success")

        # Process regular websites
        for r_idx, result in enumerate(websites):
            if len(all_results) >= num_leads:
                break
            domain = result["domain"]
            if domain in seen_domains:
                continue
            seen_domains.add(domain)

            _add_log(f"  [{r_idx+1}/{len(websites)}] Visiting {domain}...")
            contacts = extractor.fetch_and_extract(result["url"])
            lead = _build_lead_from_website(result, contacts, keyword, city, country, scorer)

            bname = lead["Business Name"].lower().strip()
            if bname not in seen_names:
                seen_names.add(bname)
                all_results.append(lead)
                if contacts["emails"]:
                    _add_log(f"    Email found: {contacts['emails'][0]}", "success")
                if contacts["phones"]:
                    _add_log(f"    Phone found: {contacts['phones'][0]}", "success")

        # Process Facebook pages
        for fb_result in fb_pages:
            if len(all_results) >= num_leads:
                break

            _add_log(f"  Extracting from Facebook: {fb_result['title'][:50]}...")
            fb_data = fb_extractor.extract(fb_result["url"], fb_result["title"], fb_result["snippet"])
            lead = _build_lead_from_facebook(fb_data, keyword, city, country, scorer)

            bname = lead["Business Name"].lower().strip()
            if bname and bname not in seen_names and bname != "unknown":
                seen_names.add(bname)
                all_results.append(lead)
                if fb_data.get("email"):
                    _add_log(f"    FB Email: {fb_data['email']}", "success")
                if fb_data.get("phone"):
                    _add_log(f"    FB Phone: {fb_data['phone']}", "success")
                if fb_data.get("website"):
                    _add_log(f"    FB Website: {fb_data['website']}", "success")

        # Process LinkedIn pages
        for li_result in li_pages:
            if len(all_results) >= num_leads:
                break

            _add_log(f"  Extracting from LinkedIn: {li_result['title'][:50]}...")
            li_data = li_extractor.extract(li_result["url"], li_result["title"], li_result["snippet"])
            lead = _build_lead_from_linkedin(li_data, keyword, city, country, scorer)

            bname = lead["Business Name"].lower().strip()
            if bname and bname not in seen_names and bname != "unknown":
                seen_names.add(bname)
                all_results.append(lead)
                if li_data.get("email"):
                    _add_log(f"    LI Email: {li_data['email']}", "success")
                if li_data.get("website"):
                    _add_log(f"    LI Website: {li_data['website']}", "success")

        # ---- GOOGLE PLACES API ----
        if places_scraper:
            _add_log(f"  Querying Google Places API for \"{keyword}\"...")
            places_query = f"{keyword} in {city}, {country}"
            places_per_kw = max(5, min(20, (num_leads - len(all_results)) // max(1, total_keywords - kw_idx + 1)))

            try:
                places_results = places_scraper.search(places_query, max_results=places_per_kw)
                total_api_calls += len(places_results) + 1  # 1 search + N detail calls

                _add_log(f"  Google Places: {len(places_results)} businesses found", "success")

                for p_idx, places_data in enumerate(places_results):
                    if len(all_results) >= num_leads:
                        break

                    bname = places_data["business_name"].lower().strip()
                    if bname in seen_names:
                        continue

                    _add_log(f"  [Places {p_idx+1}/{len(places_results)}] {places_data['business_name']}...")
                    lead = _build_lead_from_places(places_data, keyword, city, country, extractor, scorer)

                    seen_names.add(bname)
                    all_results.append(lead)

                    if lead.get("email") and lead["email"] != "N/A":
                        _add_log(f"    Places Email: {lead['email']}", "success")
                    if lead.get("phone") and lead["phone"] != "N/A":
                        _add_log(f"    Places Phone: {lead['Phone Number']}", "success")

            except Exception as e:
                _add_log(f"  Places API error: {str(e)[:100]}", "error")

        _add_log(f"  Keyword \"{keyword}\" complete -- Total leads: {len(all_results)}")

    # Sort by lead score descending
    all_results.sort(key=lambda x: x.get("Lead Score", 0), reverse=True)
    all_results = all_results[:num_leads]
    all_leads = all_results

    return {
        "leads": all_results,
        "total_leads": len(all_results),
        "api_calls": total_api_calls,
        "keywords_searched": total_keywords,
    }


# ============================================================
# API ENDPOINTS
# ============================================================

@app.post("/api/search")
async def search_leads(
    country: str = Query(...),
    city: str = Query(...),
    keywords: str = Query(""),
    num_leads: int = Query(50, ge=1, le=500),
    use_places: str = Query("false"),
    api_key: str = Query(""),
):
    """Start a multi-source lead generation search."""
    global all_leads, search_history, search_logs

    search_logs = []

    if keywords.strip():
        keywords_list = [k.strip() for k in keywords.split(",") if k.strip()][:15]
    else:
        keywords_list = ["businesses", "services", "companies"]

    use_places_api = use_places.lower() in ("true", "1", "yes")

    _add_log(f"Starting lead generation search")
    _add_log(f"Location: {city}, {country}")
    _add_log(f"Keywords: {', '.join(keywords_list)}")
    _add_log(f"Target: {num_leads} leads")
    sources = ["Google Search", "Facebook Pages", "LinkedIn Pages"]
    if use_places_api and api_key:
        sources.append("Google Places API")
    _add_log(f"Sources: {', '.join(sources)}")

    search_id = hashlib.md5(f"{city}{country}{datetime.now()}".encode()).hexdigest()[:8]

    try:
        result = await asyncio.to_thread(
            run_search, city, country, keywords_list, num_leads,
            use_places_api=use_places_api, places_api_key=api_key
        )

        search_history.insert(0, {
            "search_id": search_id,
            "city": city,
            "country": country,
            "keywords": keywords if keywords else "All categories",
            "keywords_count": len(keywords_list),
            "leads_found": result["total_leads"],
            "api_calls": result["api_calls"],
            "sources": sources,
            "timestamp": datetime.now().isoformat(),
        })
        if len(search_history) > 20:
            search_history.pop()

        _add_log(f"Search complete! Found {result['total_leads']} leads", "success")

        return {
            "success": True,
            "search_id": search_id,
            "total_leads": result["total_leads"],
            "api_calls": result["api_calls"],
            "leads": result["leads"],
            "message": f"Found {result['total_leads']} businesses in {city}, {country}",
        }

    except Exception as e:
        _add_log(f"Search failed: {str(e)}", "error")
        return JSONResponse(status_code=500, content={"error": str(e), "success": False})


@app.get("/api/leads")
async def get_leads():
    return {"leads": all_leads, "total": len(all_leads)}


@app.get("/api/stats")
async def get_stats():
    if not all_leads:
        return {"total": 0, "hot": 0, "warm": 0, "cold": 0,
                "has_email": 0, "has_phone": 0, "has_website": 0,
                "categories": {}, "avg_score": 0, "sources": {}}

    hot = len([l for l in all_leads if l.get("Lead Quality") == "Hot"])
    warm = len([l for l in all_leads if l.get("Lead Quality") == "Warm"])
    cold = len([l for l in all_leads if l.get("Lead Quality") == "Cold"])
    has_email = len([l for l in all_leads if l.get("Has Email") == "Yes"])
    has_phone = len([l for l in all_leads if l.get("Has Phone") == "Yes"])
    has_website = len([l for l in all_leads if l.get("Has Website") == "Yes"])

    categories = defaultdict(int)
    sources = defaultdict(int)
    for lead in all_leads:
        categories[lead.get("Category", "Unknown")] += 1
        sources[lead.get("Source", "Unknown")] += 1

    avg_score = sum(l.get("Lead Score", 0) for l in all_leads) // len(all_leads) if all_leads else 0

    return {
        "total": len(all_leads), "hot": hot, "warm": warm, "cold": cold,
        "has_email": has_email, "has_phone": has_phone, "has_website": has_website,
        "categories": dict(categories), "sources": dict(sources), "avg_score": avg_score,
    }


@app.get("/api/history")
async def get_history():
    return {"history": search_history}


@app.get("/api/logs")
async def get_logs():
    return {"logs": search_logs}


@app.get("/api/export")
async def export_csv():
    if not all_leads:
        return Response(content="No leads to export.", status_code=400)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"leads_export_{timestamp}.csv"
    filepath = os.path.join(os.getcwd(), filename)

    export_fields = [
        "Business Name", "Email", "All Emails", "Phone Number", "All Phones",
        "Website", "Domain", "Address", "Category", "Lead Score", "Lead Quality",
        "Facebook URL", "Instagram URL", "LinkedIn URL", "Twitter URL",
        "Has Email", "Has Phone", "Website Accessible", "Source", "Rating", "Reviews",
        "Keyword", "Extracted At",
    ]

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=export_fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(all_leads)

    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()

    return Response(content=content, media_type="text/csv",
                   headers={"Content-Disposition": f"attachment; filename={filename}"})


@app.get("/api/export/excel")
async def export_excel():
    if not all_leads:
        return Response(content="No leads to export", status_code=400)

    try:
        import pandas as pd
        from openpyxl.styles import PatternFill, Font, Alignment

        export_fields = [
            "Business Name", "Email", "All Emails", "Phone Number", "All Phones",
            "Website", "Domain", "Address", "Category", "Lead Score", "Lead Quality",
            "Facebook URL", "Instagram URL", "LinkedIn URL", "Twitter URL",
            "Has Email", "Has Phone", "Website Accessible", "Source", "Rating", "Reviews",
            "Snippet", "Keyword", "Extracted At",
        ]

        df = pd.DataFrame(all_leads)
        existing_fields = [f for f in export_fields if f in df.columns]
        df = df[existing_fields]

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"leads_export_{timestamp}.xlsx"
        filepath = os.path.join(os.getcwd(), filename)

        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Leads", index=False)
            wb = writer.book
            ws = writer.sheets["Leads"]

            header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            for col in range(1, len(existing_fields) + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for col_cells in ws.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    try:
                        if cell.value and len(str(cell.value)) > max_len:
                            max_len = len(str(cell.value))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        with open(filepath, "rb") as f:
            content = f.read()

        return Response(content=content,
                       media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       headers={"Content-Disposition": f"attachment; filename={filename}"})

    except ImportError:
        return Response(content="Install pandas and openpyxl for Excel export", status_code=500)


# ============================================================
# HTML DASHBOARD
# ============================================================
HTML_DASHBOARD = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>LeadGen Pro | Multi-Source Lead Discovery Engine</title>
    <meta name="description" content="Generate high-quality business leads from Google Search, Google Places API, Facebook and LinkedIn with email, phone, and contact extraction.">
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
    <style>
        :root {
            --bg-primary: #0a0a1a;
            --bg-secondary: #111127;
            --bg-card: rgba(20, 20, 45, 0.85);
            --glass-border: rgba(255, 255, 255, 0.08);
            --glass-shadow: 0 8px 32px rgba(0, 0, 0, 0.4);
            --text-primary: #e8e8f0;
            --text-secondary: #9090b0;
            --text-muted: #606080;
            --accent-1: #6c5ce7;
            --accent-2: #a855f7;
            --accent-3: #ec4899;
            --gradient-primary: linear-gradient(135deg, #6c5ce7, #a855f7, #ec4899);
            --gradient-success: linear-gradient(135deg, #10b981, #34d399);
            --hot: #ef4444; --warm: #f59e0b; --cold: #3b82f6;
            --success: #10b981; --warning: #f59e0b; --error: #ef4444;
            --radius: 16px; --radius-sm: 10px;
        }
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
            background: linear-gradient(135deg, #0a0a1a 0%, #111133 50%, #0a0a1a 100%);
            color: var(--text-primary); min-height: 100vh; overflow-x: hidden;
        }
        body::before, body::after {
            content: ''; position: fixed; border-radius: 50%; filter: blur(100px);
            opacity: 0.15; z-index: 0; pointer-events: none;
        }
        body::before { width: 600px; height: 600px; background: var(--accent-1); top: -200px; right: -200px; animation: float1 20s ease-in-out infinite; }
        body::after { width: 500px; height: 500px; background: var(--accent-3); bottom: -150px; left: -150px; animation: float2 25s ease-in-out infinite; }
        @keyframes float1 { 0%, 100% { transform: translate(0,0); } 50% { transform: translate(-80px,80px); } }
        @keyframes float2 { 0%, 100% { transform: translate(0,0); } 50% { transform: translate(60px,-60px); } }

        .container { max-width: 1440px; margin: 0 auto; padding: 24px; position: relative; z-index: 1; }

        .header {
            background: var(--bg-card); backdrop-filter: blur(20px); border: 1px solid var(--glass-border);
            border-radius: var(--radius); padding: 28px 32px; margin-bottom: 24px;
            box-shadow: var(--glass-shadow); display: flex; justify-content: space-between; align-items: center;
        }
        .header-left h1 { font-size: 26px; font-weight: 800; background: var(--gradient-primary); -webkit-background-clip: text; -webkit-text-fill-color: transparent; background-clip: text; }
        .header-left p { color: var(--text-secondary); font-size: 13px; margin-top: 4px; }
        .version-badge { background: rgba(108,92,231,0.15); color: var(--accent-1); padding: 6px 14px; border-radius: 20px; font-size: 12px; font-weight: 600; border: 1px solid rgba(108,92,231,0.3); }

        .search-card {
            background: var(--bg-card); backdrop-filter: blur(20px); border: 1px solid var(--glass-border);
            border-radius: var(--radius); padding: 28px 32px; margin-bottom: 24px; box-shadow: var(--glass-shadow);
        }
        .search-card h2 { font-size: 16px; font-weight: 700; margin-bottom: 20px; display: flex; align-items: center; gap: 8px; }

        .form-grid { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 16px; margin-bottom: 20px; }
        .form-group label { display: block; font-size: 12px; font-weight: 600; color: var(--text-secondary); margin-bottom: 8px; text-transform: uppercase; letter-spacing: 0.5px; }
        .form-group input, .form-group select {
            width: 100%; padding: 12px 16px; background: rgba(255,255,255,0.04); border: 1px solid var(--glass-border);
            border-radius: var(--radius-sm); color: var(--text-primary); font-size: 14px; font-family: inherit; transition: all 0.3s ease;
        }
        .form-group input:focus, .form-group select:focus { outline: none; border-color: var(--accent-1); box-shadow: 0 0 0 3px rgba(108,92,231,0.15); }
        .form-group input::placeholder { color: var(--text-muted); }
        .form-group select option { background: var(--bg-secondary); color: var(--text-primary); }

        /* Google Places toggle */
        .sources-section { margin-bottom: 20px; padding: 16px 20px; background: rgba(255,255,255,0.02); border: 1px solid var(--glass-border); border-radius: var(--radius-sm); }
        .sources-section h3 { font-size: 12px; font-weight: 600; color: var(--text-secondary); text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 14px; }
        .source-row { display: flex; align-items: center; gap: 16px; flex-wrap: wrap; }
        .source-badge { display: inline-flex; align-items: center; gap: 6px; padding: 6px 14px; border-radius: 20px; font-size: 12px; font-weight: 600; }
        .source-badge.active { background: rgba(16,185,129,0.15); color: #6ee7b7; border: 1px solid rgba(16,185,129,0.3); }
        .source-badge.optional { background: rgba(245,158,11,0.1); color: #fcd34d; border: 1px solid rgba(245,158,11,0.25); cursor: pointer; }
        .source-badge.optional.enabled { background: rgba(16,185,129,0.15); color: #6ee7b7; border: 1px solid rgba(16,185,129,0.3); }

        .toggle-switch { position: relative; display: inline-block; width: 40px; height: 22px; }
        .toggle-switch input { opacity: 0; width: 0; height: 0; }
        .toggle-slider { position: absolute; cursor: pointer; top: 0; left: 0; right: 0; bottom: 0; background: rgba(255,255,255,0.1); transition: 0.3s; border-radius: 22px; }
        .toggle-slider:before { position: absolute; content: ""; height: 16px; width: 16px; left: 3px; bottom: 3px; background: white; transition: 0.3s; border-radius: 50%; }
        .toggle-switch input:checked + .toggle-slider { background: var(--accent-1); }
        .toggle-switch input:checked + .toggle-slider:before { transform: translateX(18px); }

        .api-key-input { margin-top: 12px; display: none; }
        .api-key-input.show { display: block; }
        .api-key-input input { width: 100%; max-width: 500px; padding: 10px 14px; background: rgba(255,255,255,0.04); border: 1px solid var(--glass-border); border-radius: var(--radius-sm); color: var(--text-primary); font-size: 13px; font-family: 'JetBrains Mono', monospace; }
        .api-key-input input::placeholder { color: var(--text-muted); }
        .api-key-input input:focus { outline: none; border-color: var(--accent-1); }
        .api-key-hint { font-size: 11px; color: var(--text-muted); margin-top: 6px; }

        /* Keywords */
        .keywords-section { margin-bottom: 20px; }
        .keywords-section label { display: block; font-size: 12px; font-weight: 600; color: var(--text-secondary); margin-bottom: 8px; text-transform: uppercase; letter-spacing: 0.5px; }
        .keywords-hint { font-size: 11px; color: var(--text-muted); font-weight: 400; text-transform: none; letter-spacing: 0; margin-left: 8px; }
        .keywords-input-wrapper {
            display: flex; flex-wrap: wrap; align-items: center; gap: 8px; padding: 10px 14px;
            background: rgba(255,255,255,0.04); border: 1px solid var(--glass-border); border-radius: var(--radius-sm);
            min-height: 48px; transition: all 0.3s ease; cursor: text;
        }
        .keywords-input-wrapper:focus-within { border-color: var(--accent-1); box-shadow: 0 0 0 3px rgba(108,92,231,0.15); }
        .keyword-tag {
            display: inline-flex; align-items: center; gap: 6px; padding: 5px 12px;
            background: rgba(108,92,231,0.2); border: 1px solid rgba(108,92,231,0.4);
            border-radius: 20px; font-size: 13px; color: #c4b5fd; font-weight: 500; animation: tagIn 0.2s ease;
        }
        @keyframes tagIn { from { transform: scale(0.8); opacity: 0; } to { transform: scale(1); opacity: 1; } }
        .keyword-tag .remove-tag { cursor: pointer; font-size: 14px; opacity: 0.6; transition: opacity 0.2s; }
        .keyword-tag .remove-tag:hover { opacity: 1; }
        .keywords-text-input { flex: 1; min-width: 150px; border: none; background: none; color: var(--text-primary); font-size: 14px; font-family: inherit; outline: none; }
        .keywords-text-input::placeholder { color: var(--text-muted); }

        /* Buttons */
        .button-row { display: flex; gap: 12px; flex-wrap: wrap; }
        .btn { padding: 12px 28px; border: none; border-radius: var(--radius-sm); font-size: 14px; font-weight: 600; font-family: inherit; cursor: pointer; transition: all 0.3s ease; display: inline-flex; align-items: center; gap: 8px; }
        .btn-primary { background: var(--gradient-primary); color: white; box-shadow: 0 4px 15px rgba(108,92,231,0.3); }
        .btn-primary:hover:not(:disabled) { transform: translateY(-2px); box-shadow: 0 6px 25px rgba(108,92,231,0.4); }
        .btn-success { background: var(--gradient-success); color: white; box-shadow: 0 4px 15px rgba(16,185,129,0.3); }
        .btn-success:hover:not(:disabled) { transform: translateY(-2px); }
        .btn-outline { background: transparent; color: var(--text-secondary); border: 1px solid var(--glass-border); }
        .btn-outline:hover:not(:disabled) { background: rgba(255,255,255,0.05); color: var(--text-primary); }
        .btn:disabled { opacity: 0.5; cursor: not-allowed; transform: none !important; }

        /* Progress */
        .progress-container { margin-top: 20px; display: none; }
        .progress-bar { width: 100%; height: 6px; background: rgba(255,255,255,0.06); border-radius: 3px; overflow: hidden; }
        .progress-fill { height: 100%; background: var(--gradient-primary); width: 0%; border-radius: 3px; transition: width 0.4s ease; }
        .progress-text { font-size: 12px; color: var(--text-secondary); margin-top: 8px; text-align: center; }

        /* Stats */
        .stats-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(140px, 1fr)); gap: 16px; margin-bottom: 24px; }
        .stat-card {
            background: var(--bg-card); backdrop-filter: blur(20px); border: 1px solid var(--glass-border);
            border-radius: var(--radius); padding: 20px 18px; text-align: center; box-shadow: var(--glass-shadow); transition: all 0.3s ease;
        }
        .stat-card:hover { transform: translateY(-4px); }
        .stat-icon { font-size: 20px; margin-bottom: 6px; }
        .stat-number { font-size: 28px; font-weight: 800; letter-spacing: -1px; line-height: 1; }
        .stat-label { font-size: 10px; font-weight: 600; color: var(--text-secondary); margin-top: 6px; text-transform: uppercase; letter-spacing: 0.5px; }
        .stat-hot .stat-number { color: var(--hot); }
        .stat-warm .stat-number { color: var(--warm); }
        .stat-cold .stat-number { color: var(--cold); }
        .stat-email .stat-number { color: #a78bfa; }
        .stat-phone .stat-number { color: #34d399; }
        .stat-web .stat-number { color: #38bdf8; }

        /* Results Table */
        .results-card {
            background: var(--bg-card); backdrop-filter: blur(20px); border: 1px solid var(--glass-border);
            border-radius: var(--radius); padding: 28px 32px; margin-bottom: 24px; box-shadow: var(--glass-shadow); display: none;
        }
        .results-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 16px; }
        .results-header h3 { font-size: 16px; font-weight: 700; }
        .results-count { font-size: 12px; color: var(--text-muted); }
        .table-wrapper { overflow-x: auto; border-radius: var(--radius-sm); }
        table { width: 100%; border-collapse: collapse; font-size: 13px; }
        thead th { background: rgba(108,92,231,0.1); padding: 12px 14px; text-align: left; font-weight: 600; font-size: 11px; text-transform: uppercase; letter-spacing: 0.5px; color: var(--text-secondary); border-bottom: 1px solid var(--glass-border); position: sticky; top: 0; white-space: nowrap; }
        tbody td { padding: 10px 14px; border-bottom: 1px solid rgba(255,255,255,0.03); max-width: 200px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
        tbody tr { transition: background 0.2s; }
        tbody tr:hover { background: rgba(108,92,231,0.06); }

        .badge { display: inline-block; padding: 4px 10px; border-radius: 20px; font-size: 11px; font-weight: 700; text-transform: uppercase; }
        .badge-hot { background: rgba(239,68,68,0.15); color: #fca5a5; border: 1px solid rgba(239,68,68,0.3); }
        .badge-warm { background: rgba(245,158,11,0.15); color: #fcd34d; border: 1px solid rgba(245,158,11,0.3); }
        .badge-cold { background: rgba(59,130,246,0.15); color: #93c5fd; border: 1px solid rgba(59,130,246,0.3); }

        .score-pill { display: inline-block; padding: 3px 10px; border-radius: 20px; font-size: 12px; font-weight: 700; }
        .score-high { background: rgba(16,185,129,0.15); color: #6ee7b7; }
        .score-mid { background: rgba(245,158,11,0.15); color: #fcd34d; }
        .score-low { background: rgba(239,68,68,0.15); color: #fca5a5; }

        .source-tag { display: inline-block; padding: 3px 8px; border-radius: 12px; font-size: 10px; font-weight: 600; text-transform: uppercase; }
        .source-google { background: rgba(66,133,244,0.15); color: #93bbfc; }
        .source-facebook { background: rgba(24,119,242,0.15); color: #6db3f2; }
        .source-linkedin { background: rgba(10,102,194,0.15); color: #70b0e0; }
        .source-places { background: rgba(52,168,83,0.15); color: #6ecf8a; }

        .link-icon { color: var(--accent-2); text-decoration: none; font-weight: 600; }
        .link-icon:hover { text-decoration: underline; }
        .contact-found { color: var(--success); font-weight: 600; }
        .contact-missing { color: var(--text-muted); }

        .social-icons { display: flex; gap: 5px; }
        .social-dot { width: 8px; height: 8px; border-radius: 50%; display: inline-block; }
        .social-fb { background: #1877f2; }
        .social-ig { background: #e4405f; }
        .social-li { background: #0a66c2; }
        .social-tw { background: #1da1f2; }

        /* Logs */
        .logs-card { background: rgba(10,10,20,0.9); backdrop-filter: blur(20px); border: 1px solid var(--glass-border); border-radius: var(--radius); padding: 24px 28px; box-shadow: var(--glass-shadow); }
        .logs-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 12px; }
        .logs-header h3 { font-size: 14px; font-weight: 600; color: var(--text-secondary); }
        .logs-content { font-family: 'JetBrains Mono', 'Fira Code', 'Courier New', monospace; font-size: 12px; line-height: 1.8; max-height: 240px; overflow-y: auto; color: var(--success); }
        .logs-content::-webkit-scrollbar { width: 6px; }
        .logs-content::-webkit-scrollbar-thumb { background: rgba(255,255,255,0.1); border-radius: 3px; }
        .log-entry { padding: 3px 0; border-bottom: 1px solid rgba(255,255,255,0.03); }
        .log-time { color: var(--text-muted); margin-right: 8px; }
        .log-warning { color: var(--warning); }
        .log-error { color: var(--error); }
        .log-success { color: var(--success); }
        .log-info { color: var(--text-secondary); }

        .spinner { display: inline-block; width: 18px; height: 18px; border: 2px solid rgba(255,255,255,0.3); border-top-color: white; border-radius: 50%; animation: spin 0.7s linear infinite; }
        @keyframes spin { to { transform: rotate(360deg); } }

        @media (max-width: 900px) { .form-grid { grid-template-columns: 1fr 1fr; } .header { flex-direction: column; gap: 12px; align-items: flex-start; } }
        @media (max-width: 600px) { .container { padding: 12px; } .form-grid { grid-template-columns: 1fr; } .stats-grid { grid-template-columns: repeat(2, 1fr); } }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div class="header-left">
                <h1>LeadGen Pro</h1>
                <p>Multi-Source Lead Discovery Engine &mdash; Google Search + Places API + Facebook + LinkedIn</p>
            </div>
            <div class="version-badge">v4.0 Multi-Source</div>
        </div>

        <div class="search-card">
            <h2>New Search</h2>
            <div class="form-grid">
                <div class="form-group">
                    <label>Country</label>
                    <select id="country">
                        <option value="USA">United States</option>
                        <option value="Canada">Canada</option>
                        <option value="UK">United Kingdom</option>
                        <option value="Australia">Australia</option>
                        <option value="Germany">Germany</option>
                        <option value="France">France</option>
                        <option value="India">India</option>
                        <option value="Pakistan">Pakistan</option>
                        <option value="UAE">United Arab Emirates</option>
                        <option value="Saudi Arabia">Saudi Arabia</option>
                        <option value="South Africa">South Africa</option>
                        <option value="Brazil">Brazil</option>
                        <option value="Mexico">Mexico</option>
                        <option value="Spain">Spain</option>
                        <option value="Italy">Italy</option>
                        <option value="Netherlands">Netherlands</option>
                        <option value="Singapore">Singapore</option>
                        <option value="Japan">Japan</option>
                        <option value="New Zealand">New Zealand</option>
                    </select>
                </div>
                <div class="form-group">
                    <label>City / Area</label>
                    <input type="text" id="city" placeholder="e.g. New York, London" value="">
                </div>
                <div class="form-group">
                    <label>Max Leads</label>
                    <select id="num_leads">
                        <option value="20">20 leads</option>
                        <option value="50" selected>50 leads</option>
                        <option value="100">100 leads</option>
                        <option value="200">200 leads</option>
                        <option value="300">300 leads</option>
                        <option value="500">500 leads</option>
                    </select>
                </div>
            </div>

            <!-- Data Sources -->
            <div class="sources-section">
                <h3>Data Sources</h3>
                <div class="source-row">
                    <span class="source-badge active">Google Search</span>
                    <span class="source-badge active">Facebook Pages</span>
                    <span class="source-badge active">LinkedIn Pages</span>
                    <span class="source-badge optional" id="placesToggleBadge" onclick="togglePlaces(event)">
                        <label class="toggle-switch" style="pointer-events: none;">
                            <input type="checkbox" id="usePlaces" onchange="togglePlacesUI()">
                            <span class="toggle-slider"></span>
                        </label>
                        <span style="user-select: none;">Google Places API</span>
                    </span>
                </div>
                <div class="api-key-input" id="apiKeySection">
                    <input type="text" id="placesApiKey" placeholder="Enter your Google Places API key...">
                    <div class="api-key-hint">Get a key from <a href="https://console.cloud.google.com/apis/credentials" target="_blank" style="color:var(--accent-2);">Google Cloud Console</a> &mdash; Enable Places API</div>
                </div>
            </div>

            <!-- Keywords -->
            <div class="keywords-section">
                <label>Keywords <span class="keywords-hint">Press Enter or comma to add &middot; Multiple keywords = broader search</span></label>
                <div class="keywords-input-wrapper" id="keywordsWrapper" onclick="document.getElementById('keywordInput').focus()">
                    <input type="text" class="keywords-text-input" id="keywordInput" placeholder="Type a keyword and press Enter (e.g. plumbers, dentists, restaurants)">
                </div>
            </div>

            <div class="button-row">
                <button class="btn btn-primary" id="searchBtn" onclick="startSearch()">Generate Leads</button>
                <button class="btn btn-success" id="csvBtn" onclick="exportCSV()" disabled>Export CSV</button>
                <button class="btn btn-outline" id="excelBtn" onclick="exportExcel()" disabled>Export Excel</button>
                <button class="btn btn-outline" onclick="clearAll()">Clear</button>
            </div>

            <div class="progress-container" id="progressContainer">
                <div class="progress-bar"><div class="progress-fill" id="progressFill"></div></div>
                <div class="progress-text" id="progressText">Initializing search...</div>
            </div>
        </div>

        <div id="statsSection" style="display:none;">
            <div class="stats-grid" id="statsGrid"></div>
        </div>

        <div class="results-card" id="resultsCard">
            <div class="results-header">
                <h3>Discovered Leads</h3>
                <span class="results-count" id="resultsCount"></span>
            </div>
            <div class="table-wrapper" id="resultsTable"></div>
        </div>

        <div class="logs-card">
            <div class="logs-header">
                <h3>Live Activity Log</h3>
                <span style="font-size:11px;color:var(--text-muted);" id="logCount">0 entries</span>
            </div>
            <div class="logs-content" id="logsContent"></div>
        </div>
    </div>

    <script>
        let keywords = [];
        let currentLeads = [];
        let searchActive = false;
        let logPollInterval = null;

        // Keyword tag input
        const keywordInput = document.getElementById('keywordInput');
        keywordInput.addEventListener('keydown', function(e) {
            if (e.key === 'Enter' || e.key === ',') { e.preventDefault(); addKeyword(this.value); this.value = ''; }
            if (e.key === 'Backspace' && this.value === '' && keywords.length > 0) removeKeyword(keywords.length - 1);
        });
        keywordInput.addEventListener('paste', function(e) { setTimeout(() => { this.value.split(',').forEach(p => addKeyword(p)); this.value = ''; }, 50); });

        function addKeyword(text) { text = text.trim().replace(/,/g, ''); if (!text || keywords.includes(text)) return; keywords.push(text); renderKeywords(); }
        function removeKeyword(idx) { keywords.splice(idx, 1); renderKeywords(); }
        function renderKeywords() {
            document.getElementById('keywordsWrapper').querySelectorAll('.keyword-tag').forEach(t => t.remove());
            keywords.forEach((kw, idx) => {
                const tag = document.createElement('span');
                tag.className = 'keyword-tag';
                tag.innerHTML = kw + ' <span class="remove-tag" onclick="removeKeyword(' + idx + ')">x</span>';
                document.getElementById('keywordsWrapper').insertBefore(tag, keywordInput);
            });
        }

        // Google Places toggle
        function togglePlaces(e) {
            const checkbox = document.getElementById('usePlaces');
            checkbox.checked = !checkbox.checked;
            togglePlacesUI();
        }
        function togglePlacesUI() {
            const checked = document.getElementById('usePlaces').checked;
            document.getElementById('apiKeySection').className = checked ? 'api-key-input show' : 'api-key-input';
            document.getElementById('placesToggleBadge').className = checked ? 'source-badge optional enabled' : 'source-badge optional';
        }

        function addLog(message, level = 'info') {
            const logsDiv = document.getElementById('logsContent');
            const time = new Date().toLocaleTimeString();
            logsDiv.innerHTML += '<div class="log-entry log-' + level + '"><span class="log-time">[' + time + ']</span> ' + message + '</div>';
            logsDiv.scrollTop = logsDiv.scrollHeight;
            document.getElementById('logCount').textContent = logsDiv.children.length + ' entries';
        }

        async function pollLogs() {
            try {
                const resp = await fetch('/api/logs');
                const data = await resp.json();
                const logsDiv = document.getElementById('logsContent');
                const currentCount = logsDiv.children.length;
                if (data.logs.length > currentCount) {
                    for (let i = currentCount; i < data.logs.length; i++) {
                        const log = data.logs[i];
                        logsDiv.innerHTML += '<div class="log-entry log-' + log.level + '"><span class="log-time">[' + log.timestamp + ']</span> ' + log.message + '</div>';
                    }
                    logsDiv.scrollTop = logsDiv.scrollHeight;
                    document.getElementById('logCount').textContent = data.logs.length + ' entries';
                }
            } catch (e) {}
        }

        async function startSearch() {
            if (searchActive) return;
            const city = document.getElementById('city').value.trim();
            const country = document.getElementById('country').value;
            const numLeads = document.getElementById('num_leads').value;
            const usePlaces = document.getElementById('usePlaces').checked;
            const apiKey = document.getElementById('placesApiKey').value.trim();

            if (!city) { addLog('Please enter a city name', 'error'); return; }
            if (usePlaces && !apiKey) { addLog('Please enter your Google Places API key', 'error'); return; }

            const keywordsStr = keywords.length > 0 ? keywords.join(', ') : '';

            searchActive = true;
            const btn = document.getElementById('searchBtn');
            btn.disabled = true;
            btn.innerHTML = '<span class="spinner"></span> Searching...';
            document.getElementById('csvBtn').disabled = true;
            document.getElementById('excelBtn').disabled = true;
            document.getElementById('progressContainer').style.display = 'block';
            document.getElementById('resultsCard').style.display = 'none';
            document.getElementById('statsSection').style.display = 'none';
            document.getElementById('logsContent').innerHTML = '';

            let progress = 0;
            const progInterval = setInterval(() => {
                if (progress < 70) {
                    progress += 3;
                } else if (progress < 90) {
                    progress += 1;
                } else if (progress < 98) {
                    progress += 0.3;
                }
                progress = Math.min(progress, 98);
                document.getElementById('progressFill').style.width = progress + '%';
                document.getElementById('progressText').textContent = 'Searching and extracting contacts... ' + Math.round(progress) + '%';
            }, 500);

            logPollInterval = setInterval(pollLogs, 1000);

            try {
                let url = '/api/search?city=' + encodeURIComponent(city) + '&country=' + encodeURIComponent(country) +
                    '&keywords=' + encodeURIComponent(keywordsStr) + '&num_leads=' + numLeads +
                    '&use_places=' + usePlaces;
                if (usePlaces && apiKey) url += '&api_key=' + encodeURIComponent(apiKey);

                const response = await fetch(url, { method: 'POST' });
                const data = await response.json();

                clearInterval(progInterval);
                clearInterval(logPollInterval);
                await pollLogs();

                document.getElementById('progressFill').style.width = '100%';
                document.getElementById('progressText').textContent = data.success ? 'Complete! Discovered ' + data.total_leads + ' leads' : 'Search failed';

                if (data.success) {
                    currentLeads = data.leads;
                    addLog('Found ' + data.total_leads + ' leads!', 'success');
                    renderStats(data.leads);
                    renderTable(data.leads);
                    document.getElementById('statsSection').style.display = 'block';
                    document.getElementById('resultsCard').style.display = 'block';
                    document.getElementById('csvBtn').disabled = false;
                    document.getElementById('excelBtn').disabled = false;
                } else {
                    addLog('Search failed: ' + (data.error || 'Unknown error'), 'error');
                }
            } catch (err) {
                clearInterval(progInterval);
                clearInterval(logPollInterval);
                addLog('Network error: ' + err.message, 'error');
            } finally {
                btn.disabled = false;
                btn.innerHTML = 'Generate Leads';
                searchActive = false;
                setTimeout(() => { document.getElementById('progressContainer').style.display = 'none'; }, 1500);
            }
        }

        function renderStats(leads) {
            const hot = leads.filter(l => l['Lead Quality'] === 'Hot').length;
            const warm = leads.filter(l => l['Lead Quality'] === 'Warm').length;
            const cold = leads.filter(l => l['Lead Quality'] === 'Cold').length;
            const hasEmail = leads.filter(l => l['Has Email'] === 'Yes').length;
            const hasPhone = leads.filter(l => l['Has Phone'] === 'Yes').length;
            const hasWeb = leads.filter(l => l['Has Website'] === 'Yes').length;
            const avgScore = leads.length > 0 ? Math.round(leads.reduce((s, l) => s + (l['Lead Score'] || 0), 0) / leads.length) : 0;

            // Count sources
            const sources = {};
            leads.forEach(l => { const s = l['Source'] || 'Unknown'; sources[s] = (sources[s] || 0) + 1; });
            let sourcesHtml = '';
            Object.entries(sources).forEach(([s, c]) => {
                const cls = s.includes('Facebook') ? 'source-facebook' : s.includes('LinkedIn') ? 'source-linkedin' : s.includes('Places') ? 'source-places' : 'source-google';
                sourcesHtml += '<span class="source-tag ' + cls + '">' + s + ': ' + c + '</span> ';
            });

            document.getElementById('statsGrid').innerHTML =
                '<div class="stat-card"><div class="stat-icon">Total</div><div class="stat-number" style="background:var(--gradient-primary);-webkit-background-clip:text;-webkit-text-fill-color:transparent;">' + leads.length + '</div><div class="stat-label">Total Leads</div></div>' +
                '<div class="stat-card stat-hot"><div class="stat-icon">Hot</div><div class="stat-number">' + hot + '</div><div class="stat-label">Hot Leads</div></div>' +
                '<div class="stat-card stat-warm"><div class="stat-icon">Warm</div><div class="stat-number">' + warm + '</div><div class="stat-label">Warm Leads</div></div>' +
                '<div class="stat-card stat-cold"><div class="stat-icon">Cool</div><div class="stat-number">' + cold + '</div><div class="stat-label">Cold Leads</div></div>' +
                '<div class="stat-card stat-email"><div class="stat-icon">Email</div><div class="stat-number">' + hasEmail + '</div><div class="stat-label">Have Email</div></div>' +
                '<div class="stat-card stat-phone"><div class="stat-icon">Phone</div><div class="stat-number">' + hasPhone + '</div><div class="stat-label">Have Phone</div></div>' +
                '<div class="stat-card stat-web"><div class="stat-icon">Web</div><div class="stat-number">' + hasWeb + '</div><div class="stat-label">Have Website</div></div>' +
                '<div class="stat-card"><div class="stat-icon">Score</div><div class="stat-number" style="background:var(--gradient-primary);-webkit-background-clip:text;-webkit-text-fill-color:transparent;">' + avgScore + '</div><div class="stat-label">Avg Score</div></div>';
        }

        function renderTable(leads) {
            if (!leads || leads.length === 0) {
                document.getElementById('resultsTable').innerHTML = '<p style="color:var(--text-muted);padding:20px;">No leads found.</p>';
                return;
            }
            document.getElementById('resultsCount').textContent = 'Showing ' + Math.min(leads.length, 100) + ' of ' + leads.length + ' leads';

            let html = '<table><thead><tr><th>#</th><th>Business</th><th>Email</th><th>Phone</th><th>Website</th><th>Category</th><th>Score</th><th>Quality</th><th>Source</th><th>Socials</th></tr></thead><tbody>';

            leads.slice(0, 100).forEach((lead, idx) => {
                const qualityClass = 'badge-' + (lead['Lead Quality'] || 'cold').toLowerCase();
                const score = lead['Lead Score'] || 0;
                const scoreClass = score >= 70 ? 'score-high' : (score >= 40 ? 'score-mid' : 'score-low');
                const email = lead['Email'] && lead['Email'] !== 'N/A' ? '<span class="contact-found">' + lead['Email'] + '</span>' : '<span class="contact-missing">-</span>';
                const phone = lead['Phone Number'] && lead['Phone Number'] !== 'N/A' ? '<span class="contact-found">' + lead['Phone Number'] + '</span>' : '<span class="contact-missing">-</span>';
                const website = lead['Website'] ? '<a href="' + lead['Website'] + '" target="_blank" class="link-icon">' + (lead['Domain'] || '').substring(0, 25) + '</a>' : '-';

                const src = lead['Source'] || '';
                const srcClass = src.includes('Facebook') ? 'source-facebook' : src.includes('LinkedIn') ? 'source-linkedin' : src.includes('Places') ? 'source-places' : 'source-google';

                let socials = '<div class="social-icons">';
                if (lead['Facebook URL']) socials += '<span class="social-dot social-fb" title="Facebook"></span>';
                if (lead['Instagram URL']) socials += '<span class="social-dot social-ig" title="Instagram"></span>';
                if (lead['LinkedIn URL']) socials += '<span class="social-dot social-li" title="LinkedIn"></span>';
                if (lead['Twitter URL']) socials += '<span class="social-dot social-tw" title="Twitter"></span>';
                socials += '</div>';

                html += '<tr>' +
                    '<td>' + (idx + 1) + '</td>' +
                    '<td><strong>' + (lead['Business Name'] || '-') + '</strong></td>' +
                    '<td>' + email + '</td>' +
                    '<td>' + phone + '</td>' +
                    '<td>' + website + '</td>' +
                    '<td>' + (lead['Category'] || '-') + '</td>' +
                    '<td><span class="score-pill ' + scoreClass + '">' + score + '</span></td>' +
                    '<td><span class="badge ' + qualityClass + '">' + (lead['Lead Quality'] || '-') + '</span></td>' +
                    '<td><span class="source-tag ' + srcClass + '">' + src + '</span></td>' +
                    '<td>' + socials + '</td>' +
                    '</tr>';
            });

            html += '</tbody></table>';
            if (leads.length > 100) html += '<p style="text-align:center;margin-top:16px;color:var(--text-muted);font-size:13px;">Showing 100 of ' + leads.length + ' leads. Export to see all.</p>';
            document.getElementById('resultsTable').innerHTML = html;
        }

        function exportCSV() { if (!currentLeads.length) { addLog('No leads', 'error'); return; } addLog('Downloading CSV...'); window.location.href = '/api/export'; }
        function exportExcel() { if (!currentLeads.length) { addLog('No leads', 'error'); return; } addLog('Downloading Excel...'); window.location.href = '/api/export/excel'; }
        function clearAll() {
            currentLeads = []; keywords = []; renderKeywords();
            document.getElementById('resultsCard').style.display = 'none';
            document.getElementById('statsSection').style.display = 'none';
            document.getElementById('logsContent').innerHTML = '';
            document.getElementById('csvBtn').disabled = true;
            document.getElementById('excelBtn').disabled = true;
            addLog('Cleared all data');
        }

        addLog('LeadGen Pro v4.0 - Ready');
        addLog('Sources: Google Search + Facebook + LinkedIn (always on)');
        addLog('Optional: Enable Google Places API for structured business data');
    </script>
</body>
</html>
"""


@app.get("/")
@app.get("/dashboard")
async def dashboard():
    return HTMLResponse(HTML_DASHBOARD)


# ============================================================
# MAIN
# ============================================================
if __name__ == "__main__":
    import sys
    if sys.platform == "win32":
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
            sys.stderr.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass

    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    print("\n" + "=" * 70)
    print("  LEADGEN PRO v4.0 - Multi-Source Lead Generation Platform")
    print("=" * 70)
    print("\n  SOURCES:")
    print("   - Google Search scraping (always on, no API key)")
    print("   - Facebook page extraction (automatic)")
    print("   - LinkedIn company extraction (automatic)")
    print("   - Google Places API (optional, needs API key)")
    print("\n  FEATURES:")
    print("   - Extract emails, phone numbers & social links")
    print("   - Multi-keyword search (comma-separated)")
    print("   - Lead scoring (0-100) with Hot/Warm/Cold tiers")
    print("   - Export to CSV or Excel")
    print(f"   - Files save to: {os.getcwd()}")
    print("\n  Dashboard: http://localhost:8000")
    print("\n" + "=" * 70 + "\n")

    uvicorn.run(app, host="0.0.0.0", port=8000, log_level="info")